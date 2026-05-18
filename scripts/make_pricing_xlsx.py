"""Build the LLM API pricing estimate workbook.

Single-shot script: produces LLM_API_Pricing_Estimate.xlsx at the repo root.
Light professional palette (white + Bajaj blue), no dark fills.
Both USD and INR columns shown; conversion rate not exposed in the sheet.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

# Internal conversion rate (not displayed in the workbook).
USD_TO_INR = 84.0

# ---- Palette (light only, no dark colors) -----------------------------------
ACCENT      = "0072CE"   # Bajaj blue (header bar)
ACCENT_LITE = "E8F1FA"   # very light blue (banded row)
ACCENT_EDGE = "0072CE"   # accent border
HEADER_TXT  = "FFFFFF"
BODY_TXT    = "1F2937"
BORDER_CLR  = "B8C5D6"

thin   = Side(style="thin",   color=BORDER_CLR)
medium = Side(style="medium", color=ACCENT_EDGE)

box           = Border(left=thin,   right=thin,   top=thin,   bottom=thin)
box_top_md    = Border(left=medium, right=medium, top=medium, bottom=thin)
box_bottom_md = Border(left=medium, right=medium, top=thin,   bottom=medium)
box_mid_md    = Border(left=medium, right=medium, top=thin,   bottom=thin)
box_solo_md   = Border(left=medium, right=medium, top=medium, bottom=medium)

font_title    = Font(name="Calibri", size=16, bold=True, color=BODY_TXT)
font_url      = Font(name="Calibri", size=10, italic=True, color="0563C1", underline="single")
font_header   = Font(name="Calibri", size=10, bold=True, color=HEADER_TXT)
font_body     = Font(name="Calibri", size=10, color=BODY_TXT)
font_total    = Font(name="Calibri", size=11, bold=True, color=BODY_TXT)

fill_accent   = PatternFill("solid", fgColor=ACCENT)
fill_accent_l = PatternFill("solid", fgColor=ACCENT_LITE)

center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left   = Alignment(horizontal="left",   vertical="center", wrap_text=True, indent=1)
right  = Alignment(horizontal="right",  vertical="center", indent=1)

# ---- Per-provider data (verified or user-supplied prices) -------------------
PROVIDERS = [
    {
        "sheet": "Claude",
        "title": "Claude (Anthropic)",
        "url":   "https://platform.claude.com/docs/en/docs/about-claude/pricing",
        "rows":  [
            # model,                 input $/MTok, output $/MTok, mo_in_M, mo_out_M
            ("Claude Opus 4.7",      5.00,         25.00,         10.0,    3.0),
        ],
    },
    {
        "sheet": "Gemini",
        "title": "Gemini (Google)",
        "url":   "https://ai.google.dev/gemini-api/docs/pricing",
        "rows":  [
            ("Gemini 3.1 Pro Preview (<=200k)", 2.00, 12.00, 0.5, 0.3),
        ],
    },
    {
        "sheet": "GPT",
        "title": "GPT (OpenAI)",
        "url":   "https://openai.com/api/pricing/",
        "rows":  [
            ("GPT-5.5", 5.00, 30.00, 0.5, 0.3),
        ],
    },
    {
        "sheet": "Perplexity",
        "title": "Perplexity",
        "url":   "https://docs.perplexity.ai/guides/pricing",
        "rows":  [
            ("Sonar", 1.00, 1.00, 0.5, 0.3),
        ],
    },
]

COL_HEADERS = [
    "Model",
    "Input rate (USD / 1M tokens)",
    "Output rate (USD / 1M tokens)",
    "Est. monthly input tokens (M)",
    "Est. monthly output tokens (M)",
    "Input cost (USD)",
    "Output cost (USD)",
    "Total monthly (USD)",
    "Total monthly (INR)",
]

COL_WIDTHS = [36, 22, 22, 22, 22, 18, 18, 20, 22]


def style_header_row(ws, row, n_cols):
    last = n_cols
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = font_header
        cell.fill = fill_accent
        cell.alignment = center
        # outer accent border on header strip
        if c == 1 and c == last:
            cell.border = box_solo_md
        elif c == 1:
            cell.border = Border(left=medium, right=thin, top=medium, bottom=medium)
        elif c == last:
            cell.border = Border(left=thin,   right=medium, top=medium, bottom=medium)
        else:
            cell.border = Border(left=thin, right=thin, top=medium, bottom=medium)
    ws.row_dimensions[row].height = 34


def _body_border(col_idx, last_col):
    """Body cell — accent border on outer sides, thin between."""
    left_s   = medium if col_idx == 1 else thin
    right_s  = medium if col_idx == last_col else thin
    return Border(left=left_s, right=right_s, top=thin, bottom=thin)


def _total_border(col_idx, last_col):
    """Totals row — accent border on all outer + bottom."""
    left_s   = medium if col_idx == 1 else thin
    right_s  = medium if col_idx == last_col else thin
    return Border(left=left_s, right=right_s, top=thin, bottom=medium)


def write_provider_sheet(wb, p):
    ws = wb.create_sheet(p["sheet"])
    n_cols = len(COL_HEADERS)

    # Column widths
    for i, w in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Row 1: title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    t = ws.cell(row=1, column=1, value=p["title"])
    t.font = font_title
    t.alignment = left
    ws.row_dimensions[1].height = 32

    # Row 2: pricing source URL
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    u = ws.cell(row=2, column=1, value=p["url"])
    u.hyperlink = p["url"]
    u.font = font_url
    u.alignment = left
    ws.row_dimensions[2].height = 20

    # Row 3 blank spacer
    ws.row_dimensions[3].height = 10

    # Row 4: column headers
    for ci, h in enumerate(COL_HEADERS, start=1):
        ws.cell(row=4, column=ci, value=h)
    style_header_row(ws, 4, n_cols)

    # Data rows
    start_row = 5
    row = start_row
    for i, (model, in_rate, out_rate, mo_in_m, mo_out_m) in enumerate(p["rows"]):
        in_cost  = in_rate  * mo_in_m
        out_cost = out_rate * mo_out_m
        tot_usd  = in_cost + out_cost
        tot_inr  = tot_usd * USD_TO_INR

        values = [model, in_rate, out_rate, mo_in_m, mo_out_m,
                  in_cost, out_cost, tot_usd, tot_inr]
        for ci, v in enumerate(values, start=1):
            c = ws.cell(row=row, column=ci, value=v)
            c.border = _body_border(ci, n_cols)
            c.font = font_body
            c.alignment = left if ci == 1 else right
            if i % 2 == 1:
                c.fill = fill_accent_l
        # Number formats
        ws.cell(row=row, column=2).number_format = '"$"#,##0.00'
        ws.cell(row=row, column=3).number_format = '"$"#,##0.00'
        ws.cell(row=row, column=4).number_format = '#,##0.00" M"'
        ws.cell(row=row, column=5).number_format = '#,##0.00" M"'
        ws.cell(row=row, column=6).number_format = '"$"#,##0.00'
        ws.cell(row=row, column=7).number_format = '"$"#,##0.00'
        ws.cell(row=row, column=8).number_format = '"$"#,##0.00'
        ws.cell(row=row, column=9).number_format = '"₹"#,##0'
        ws.row_dimensions[row].height = 22
        row += 1

    last_data_row = row - 1

    # Totals row
    for ci in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=ci)
        cell.fill = fill_accent_l
        cell.border = _total_border(ci, n_cols)
        cell.font = font_total
        cell.alignment = left if ci == 1 else right
    ws.cell(row=row, column=1, value="Subtotal")
    usd_cell = ws.cell(row=row, column=8,
                       value=f"=SUM(H{start_row}:H{last_data_row})")
    usd_cell.number_format = '"$"#,##0.00'
    inr_cell = ws.cell(row=row, column=9,
                       value=f"=SUM(I{start_row}:I{last_data_row})")
    inr_cell.number_format = '"₹"#,##0'
    ws.row_dimensions[row].height = 26

    # Freeze top
    ws.freeze_panes = "A5"
    ws.sheet_view.showGridLines = False

    return ws.title, last_data_row, row  # name, last data row, totals row


def write_summary_sheet(wb, refs):
    ws = wb.create_sheet("Summary", 0)
    n_cols = 3
    for i, w in enumerate([34, 26, 26], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.merge_cells("A1:C1")
    t = ws.cell(row=1, column=1, value="LLM API Pricing Estimate")
    t.font = font_title
    t.alignment = left
    ws.row_dimensions[1].height = 32

    ws.row_dimensions[2].height = 10

    headers = ["Provider", "Total monthly (USD)", "Total monthly (INR)"]
    for ci, h in enumerate(headers, start=1):
        ws.cell(row=3, column=ci, value=h)
    style_header_row(ws, 3, n_cols)

    row = 4
    first = row
    for prov_name, _last_data_row, totals_row in refs:
        name_cell = ws.cell(row=row, column=1, value=prov_name)
        name_cell.font = font_body
        name_cell.alignment = left
        name_cell.border = _body_border(1, n_cols)

        usd = ws.cell(row=row, column=2, value=f"='{prov_name}'!H{totals_row}")
        usd.number_format = '"$"#,##0.00'
        usd.font = font_body
        usd.alignment = right
        usd.border = _body_border(2, n_cols)

        inr = ws.cell(row=row, column=3, value=f"='{prov_name}'!I{totals_row}")
        inr.number_format = '"₹"#,##0'
        inr.font = font_body
        inr.alignment = right
        inr.border = _body_border(3, n_cols)

        if (row - first) % 2 == 1:
            for c in range(1, n_cols + 1):
                ws.cell(row=row, column=c).fill = fill_accent_l
        ws.row_dimensions[row].height = 22
        row += 1

    last = row - 1

    # Grand total
    for ci in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=ci)
        cell.fill = fill_accent_l
        cell.border = _total_border(ci, n_cols)
        cell.font = font_total
        cell.alignment = left if ci == 1 else right
    ws.cell(row=row, column=1, value="Grand total")
    g_usd = ws.cell(row=row, column=2, value=f"=SUM(B{first}:B{last})")
    g_usd.number_format = '"$"#,##0.00'
    g_inr = ws.cell(row=row, column=3, value=f"=SUM(C{first}:C{last})")
    g_inr.number_format = '"₹"#,##0'
    ws.row_dimensions[row].height = 26

    ws.freeze_panes = "A4"
    ws.sheet_view.showGridLines = False
    ws.sheet_view.tabSelected = True


def main():
    wb = Workbook()
    # Remove default sheet so order is clean
    default = wb.active
    wb.remove(default)

    refs = []
    for p in PROVIDERS:
        name, last_data, totals_row = write_provider_sheet(wb, p)
        refs.append((name, last_data, totals_row))

    write_summary_sheet(wb, refs)
    wb.active = 0  # Summary first

    base = Path(__file__).resolve().parents[1] / "LLM_API_Pricing_Estimate.xlsx"
    out = base
    try:
        wb.save(out)
    except PermissionError:
        # File is open in Excel; write to a versioned filename instead.
        import time
        out = base.with_name(f"{base.stem}_v{int(time.time())}.xlsx")
        wb.save(out)
    print(f"Wrote {out}")


if __name__ == "__main__":
    main()
