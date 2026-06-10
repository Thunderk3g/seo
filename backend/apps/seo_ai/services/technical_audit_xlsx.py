"""XLSX exporter for the technical audit engine.

Single-URL report (multi-sheet): Summary + KPIs, Findings (drawback +
recommendation), Images-missing-alt, Links, Broken-links, CWV. Reused
for a whole-site audit (one Findings-rollup sheet across the latest
Bajaj snapshot's pages). Pure openpyxl — Bajaj-blue palette.
"""
from __future__ import annotations

import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

NAVY = "002C6E"
BLUE = "0072CE"
WHITE = "FFFFFF"
RED = "FBE9E7"
ORANGE = "FFF3E0"
GREY = "F1F5F9"

_HEAD = Font(name="Calibri", size=11, bold=True, color=WHITE)
_TITLE = Font(name="Calibri", size=18, bold=True, color=NAVY)
_LABEL = Font(name="Calibri", size=10, bold=True, color="475569")
_HEAD_FILL = PatternFill("solid", fgColor=NAVY)
_SEV_FILL = {"critical": PatternFill("solid", fgColor=RED),
             "warning": PatternFill("solid", fgColor=ORANGE),
             "notice": PatternFill("solid", fgColor=GREY)}
_WRAP = Alignment(wrap_text=True, vertical="top")


def _header_row(ws, row: int, headers: list[str]) -> None:
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = _HEAD
        cell.fill = _HEAD_FILL


def _autofit(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _summary_sheet(wb: Workbook, audit: dict[str, Any]) -> None:
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "Technical SEO Audit"
    ws["A1"].font = _TITLE
    ws["A2"] = audit.get("url", "")
    ws["A2"].font = Font(name="Calibri", size=11, color=BLUE)
    ws["A3"] = (f"Source: {audit.get('source')}  ·  Score: "
                f"{audit.get('score')}/100  ·  "
                f"{audit['counts']['critical']} critical / "
                f"{audit['counts']['warning']} warning / "
                f"{audit['counts']['notice']} notice")
    ws["A3"].font = _LABEL

    s = audit.get("summary", {})
    rows = [
        ("Title", s.get("title")), ("Title length", s.get("title_length")),
        ("Meta description length", s.get("meta_description_length")),
        ("Word count", s.get("word_count")),
        ("H1 / H2 / H3", f"{s.get('h1')} / {s.get('h2')} / {s.get('h3')}"),
        ("Internal links", s.get("internal_links")),
        ("External links", s.get("external_links")),
        ("Images total", s.get("images_total")),
        ("Images missing alt", s.get("images_missing_alt")),
        ("Schema types", ", ".join(s.get("schema_types") or [])),
        ("Canonical", s.get("canonical")),
        ("Status code", s.get("status_code")),
        ("Response time (ms)", s.get("response_time_ms")),
    ]
    cwv = audit.get("cwv") or {}
    if cwv.get("available"):
        m = (cwv.get("mobile") or {})
        lab = m.get("lab") or {}
        field = m.get("field") or {}
        rows.append(("Mobile perf score", m.get("performance_score")))
        rows.append(("Mobile LCP ms (field/lab)",
                     f"{field.get('lcp_ms') or '-'} / {lab.get('lcp_ms') or '-'}"))
        rows.append(("Mobile CLS (field/lab)",
                     f"{field.get('cls') if field.get('cls') is not None else '-'} / {lab.get('cls')}"))
        rows.append(("Field INP ms", field.get("inp_ms") or "-"))
    r0 = 5
    for i, (k, v) in enumerate(rows):
        ws.cell(row=r0 + i, column=1, value=k).font = _LABEL
        ws.cell(row=r0 + i, column=2, value=v)
    _autofit(ws, [28, 80])


def _findings_sheet(wb: Workbook, findings: list[dict]) -> None:
    ws = wb.create_sheet("Findings")
    _header_row(ws, 1, ["Severity", "Check", "Drawback", "Recommendation"])
    for i, f in enumerate(findings, 2):
        ws.cell(row=i, column=1, value=f.get("severity"))
        ws.cell(row=i, column=2, value=f.get("check"))
        ws.cell(row=i, column=3, value=f.get("detail")).alignment = _WRAP
        ws.cell(row=i, column=4, value=f.get("recommendation")).alignment = _WRAP
        fill = _SEV_FILL.get(f.get("severity"))
        if fill:
            ws.cell(row=i, column=1).fill = fill
    _autofit(ws, [12, 18, 60, 60])


def _list_sheet(wb: Workbook, title: str, headers: list[str],
                rows: list[list], widths: list[int]) -> None:
    ws = wb.create_sheet(title)
    _header_row(ws, 1, headers)
    for i, r in enumerate(rows, 2):
        for c, v in enumerate(r, 1):
            ws.cell(row=i, column=c, value=v)
    _autofit(ws, widths)


def build_single_url_xlsx(audit: dict[str, Any]) -> bytes:
    wb = Workbook()
    _summary_sheet(wb, audit)
    _findings_sheet(wb, audit.get("findings", []))

    miss = audit.get("images_missing_alt_samples", [])
    if miss:
        _list_sheet(wb, "Images missing alt", ["Image URL"],
                    [[u] for u in miss], [100])
    broken = audit.get("broken_links", [])
    if broken:
        _list_sheet(wb, "Broken links", ["URL", "Status"],
                    [[b.get("url"), b.get("status")] for b in broken], [90, 12])
    outline = audit.get("h2_outline", [])
    if outline:
        _list_sheet(wb, "H2 outline", ["H2 heading"],
                    [[h] for h in outline], [80])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_site_audit_xlsx(audit: dict[str, Any]) -> bytes:
    """Whole-site aggregate audit (output of ``audit_site``): a Findings
    sheet (issue · count · drawback · recommendation) plus one sheet per
    issue listing affected sample URLs."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Site audit"
    ws["A1"] = "Technical SEO Audit — full website"
    ws["A1"].font = _TITLE
    snap = audit.get("snapshot", {})
    ws["A2"] = (f"{snap.get('kind')} snapshot · {snap.get('pages')} pages · "
                f"score {audit.get('site_score')}/100 · "
                f"{audit['counts']['critical']} critical / "
                f"{audit['counts']['warning']} warning / "
                f"{audit['counts']['notice']} notice")
    ws["A2"].font = _LABEL

    _header_row(ws, 4, ["Severity", "Issue", "Count", "Drawback", "Recommendation"])
    for i, f in enumerate(audit.get("findings", []), 5):
        ws.cell(row=i, column=1, value=f.get("severity"))
        ws.cell(row=i, column=2, value=f.get("check"))
        ws.cell(row=i, column=3, value=f.get("count"))
        ws.cell(row=i, column=4, value=f.get("detail")).alignment = _WRAP
        ws.cell(row=i, column=5, value=f.get("recommendation")).alignment = _WRAP
        fill = _SEV_FILL.get(f.get("severity"))
        if fill:
            ws.cell(row=i, column=1).fill = fill
    _autofit(ws, [12, 20, 8, 55, 55])

    # One sheet per issue with sample affected URLs.
    for f in audit.get("findings", []):
        samples = f.get("samples") or []
        if not samples:
            continue
        name = (f.get("check") or "issue")[:28]
        ws2 = wb.create_sheet(name)
        _header_row(ws2, 1, ["Affected (sample)"])
        for j, u in enumerate(samples, 2):
            ws2.cell(row=j, column=1, value=u)
        _autofit(ws2, [100])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_site_xlsx(rows: list[dict], *, snapshot_label: str = "") -> bytes:
    """Whole-site technical rollup: one row per audited page with its
    score + issue counts. ``rows`` = list of audit_url() outputs."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Site technical audit"
    ws["A1"] = "Technical SEO Audit — full site"
    ws["A1"].font = _TITLE
    ws["A2"] = snapshot_label
    ws["A2"].font = _LABEL
    _header_row(ws, 4, ["URL", "Score", "Critical", "Warning", "Notice",
                        "Words", "H1", "Missing alt", "Int links", "LCP ms"])
    for i, a in enumerate(rows, 5):
        if not a.get("ok"):
            continue
        s = a["summary"]
        cwv = a.get("cwv") or {}
        m = (cwv.get("mobile") or {}) if cwv.get("available") else {}
        field = (m or {}).get("field") or {}
        lab = (m or {}).get("lab") or {}
        ws.cell(row=i, column=1, value=a["url"])
        ws.cell(row=i, column=2, value=a["score"])
        ws.cell(row=i, column=3, value=a["counts"]["critical"])
        ws.cell(row=i, column=4, value=a["counts"]["warning"])
        ws.cell(row=i, column=5, value=a["counts"]["notice"])
        ws.cell(row=i, column=6, value=s.get("word_count"))
        ws.cell(row=i, column=7, value=s.get("h1"))
        ws.cell(row=i, column=8, value=s.get("images_missing_alt"))
        ws.cell(row=i, column=9, value=s.get("internal_links"))
        ws.cell(row=i, column=10, value=field.get("lcp_ms") or lab.get("lcp_ms"))
    _autofit(ws, [70, 8, 9, 9, 8, 8, 6, 11, 10, 9])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
