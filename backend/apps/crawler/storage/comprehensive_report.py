"""Comprehensive XLSX report bundling Phase A-D audit data.

The legacy ``excel_writer.build_report`` only covers the original
crawl-result sheets. After Phases A-D we have 66+ new signals
(security headers, hreflang, JSON-LD, render-delta, PDFs, custom
extractors, readability, cookies, AMP, accessibility) that deserve
their own organised sheets — manager-presentable, with charts.

This module builds a separate ``crawl_report_comprehensive.xlsx``
with the following sheet layout (each sheet is omitted if its data
section is unchecked in the field selector):

  1. Executive Summary       — KPIs + compliance pie
  2. Compliance Overview     — WCAG/GDPR/OWASP totals + bar chart
  3. WCAG Findings           — per-rule rows with URL + evidence
  4. Privacy & Cookies       — cookies sheet + tracker analysis
  5. Security Headers        — per-URL header presence matrix
  6. Structured Data         — JSON-LD types per URL + rich-result eligibility
  7. Hreflang Matrix         — locale clusters with return-tag status
  8. Technical SEO           — canonical chains + redirect chains
  9. Content Audit           — titles + meta + readability + pixel widths
 10. Page Inventory          — all crawled URLs with key Phase A-D metrics
 11. Detector Catalog        — every rule + count + fix instructions

Reuses styling primitives from ``excel_writer`` (Bajaj-blue header,
zebra rows, status-code conditional fills) so this report visually
matches the legacy one.
"""
from __future__ import annotations

import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from ..conf import settings
from .excel_writer import (
    BRAND_NAVY, BRAND_GOLD, OK_GREEN, ERR_RED, TEXT_DARK, TEXT_MUTED,
    ZEBRA, WHITE,
    _HEADER_FONT, _SUBHEAD_FONT, _TITLE_FONT, _SUB_FONT,
    _KPI_LABEL, _KPI_VALUE,
    _HEADER_FILL, _ACCENT_FILL, _ZEBRA_FILL, _OK_FILL, _ERR_FILL,
    _BORDER, _CENTER, _LEFT,
    _style_header, _fit_columns,
)


# Every section the report supports. Keys correspond to the
# ``sections`` query-param on the API endpoint. Order is the sheet
# emission order; later sheets are emitted after earlier ones.
ALL_SECTIONS: tuple[str, ...] = (
    "summary",
    "compliance",
    "wcag",
    "privacy",
    "security",
    "structured_data",
    "hreflang",
    "technical",
    "content",
    "inventory",
    "catalog",
)


def _row_list(v: Any) -> list:
    if isinstance(v, list):
        return v
    if not v:
        return []
    try:
        parsed = json.loads(str(v))
        return parsed if isinstance(parsed, list) else []
    except (ValueError, TypeError):
        return []


def _row_dict(v: Any) -> dict:
    if isinstance(v, dict):
        return v
    if not v:
        return {}
    try:
        parsed = json.loads(str(v))
        return parsed if isinstance(parsed, dict) else {}
    except (ValueError, TypeError):
        return {}


def _safe_int(v: Any) -> int:
    try:
        return int(v) if v not in (None, "") else 0
    except (TypeError, ValueError):
        return 0


# ── 1. Executive Summary ──────────────────────────────────────────


def _build_summary(ws: Worksheet, audit_rows: list[dict], compliance: dict) -> None:
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:G1")
    ws["A1"] = "Bajaj Life Insurance — SEO & Compliance Report"
    ws["A1"].font = _TITLE_FONT
    ws["A1"].alignment = _LEFT
    ws.row_dimensions[1].height = 36

    ws.merge_cells("A2:G2")
    ws["A2"] = (
        "Comprehensive site audit covering technical SEO, WCAG 2.1 "
        "accessibility, GDPR/DPDPA privacy, and OWASP security headers."
    )
    ws["A2"].font = _SUB_FONT

    summary = compliance.get("summary", {}) or {}
    pages_audited = len(audit_rows)
    pages_ok = sum(1 for r in audit_rows if (r.get("status_code") or "") == "200")

    kpis = [
        ("Pages crawled", pages_audited),
        ("Pages OK (200)", pages_ok),
        ("Pages with violations", summary.get("pages_with_any_violation", 0)),
        ("Total violations", summary.get("total_violations", 0)),
        ("Unique rules failed", summary.get("unique_rules_failed", 0)),
        ("Errors", summary.get("by_severity", {}).get("error", 0)),
        ("Warnings", summary.get("by_severity", {}).get("warning", 0)),
        ("Notices", summary.get("by_severity", {}).get("notice", 0)),
    ]
    start_row = 4
    for i, (label, value) in enumerate(kpis):
        col = (i % 4) * 2 + 1
        row = start_row + (i // 4) * 3
        ws.cell(row=row, column=col, value=label).font = _KPI_LABEL
        cv = ws.cell(row=row + 1, column=col, value=value)
        cv.font = _KPI_VALUE
        cv.alignment = _LEFT
        ws.merge_cells(
            start_row=row + 1, start_column=col,
            end_row=row + 1, end_column=col + 1,
        )

    by_sect = summary.get("by_section", {}) or {}
    if any(v > 0 for v in by_sect.values()):
        chart_data_row = 12
        ws.cell(row=chart_data_row, column=1, value="Section").font = _SUBHEAD_FONT
        ws.cell(row=chart_data_row, column=2, value="Violations").font = _SUBHEAD_FONT
        section_labels = {
            "wcag": "WCAG Accessibility",
            "privacy": "Privacy & Cookies",
            "security_headers": "Security Headers",
        }
        rownum = chart_data_row + 1
        for key, label in section_labels.items():
            ws.cell(row=rownum, column=1, value=label)
            ws.cell(row=rownum, column=2, value=int(by_sect.get(key, 0)))
            rownum += 1
        chart = PieChart()
        chart.title = "Violations by compliance section"
        labels = Reference(ws, min_col=1, min_row=chart_data_row + 1,
                           max_row=rownum - 1)
        data = Reference(ws, min_col=2, min_row=chart_data_row,
                         max_row=rownum - 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(labels)
        chart.height = 9
        chart.width = 14
        chart.dataLabels = DataLabelList(showPercent=True)
        ws.add_chart(chart, "D12")

    for col in "ABCDEFGH":
        ws.column_dimensions[col].width = 22


# ── 2. Compliance Overview ────────────────────────────────────────


def _build_compliance_overview(ws: Worksheet, compliance: dict) -> None:
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:F1")
    ws["A1"] = "Compliance Overview — WCAG / GDPR / OWASP"
    ws["A1"].font = _TITLE_FONT
    ws.row_dimensions[1].height = 30

    headers = ["Section", "Rule", "Severity", "Standard", "Reference", "Count"]
    ws.append([])
    ws.append(headers)
    _style_header(ws, 3, len(headers))

    row_idx = 4
    chart_rows: list[tuple[str, int]] = []
    for section in compliance.get("sections", []):
        section_total = section.get("total_violations", 0)
        if section_total > 0:
            chart_rows.append((section["title"], section_total))
        for rule in section.get("rules", []):
            refs = rule.get("references", [])
            ref_str = "; ".join(
                f'{r.get("standard","")} {r.get("ref","")}' for r in refs
            ) if refs else ""
            standard_str = ", ".join({r.get("standard", "") for r in refs})
            ws.cell(row=row_idx, column=1, value=section.get("title", ""))
            ws.cell(row=row_idx, column=2, value=rule.get("title", ""))
            sev_cell = ws.cell(row=row_idx, column=3, value=rule.get("severity", ""))
            if rule.get("severity") == "error":
                sev_cell.font = Font(name="Calibri", size=10, bold=True, color=ERR_RED)
            elif rule.get("severity") == "warning":
                sev_cell.font = Font(name="Calibri", size=10, bold=True, color="C28E00")
            ws.cell(row=row_idx, column=4, value=standard_str)
            ws.cell(row=row_idx, column=5, value=ref_str)
            ws.cell(row=row_idx, column=6, value=rule.get("count", 0))
            for ci in range(1, 7):
                c = ws.cell(row=row_idx, column=ci)
                c.alignment = _LEFT
                c.border = _BORDER
                if row_idx % 2 == 0:
                    c.fill = _ZEBRA_FILL
            row_idx += 1

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:F{row_idx - 1}"

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 48
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 28
    ws.column_dimensions["F"].width = 10

    # Section bar chart
    if chart_rows:
        chart_start = row_idx + 2
        ws.cell(row=chart_start, column=1, value="Section").font = _SUBHEAD_FONT
        ws.cell(row=chart_start, column=2, value="Violations").font = _SUBHEAD_FONT
        for i, (label, value) in enumerate(chart_rows, start=1):
            ws.cell(row=chart_start + i, column=1, value=label)
            ws.cell(row=chart_start + i, column=2, value=int(value))
        chart = BarChart()
        chart.type = "bar"
        chart.title = "Violations by section"
        chart.y_axis.title = "Section"
        chart.x_axis.title = "Violation count"
        data = Reference(ws, min_col=2, min_row=chart_start,
                         max_row=chart_start + len(chart_rows))
        cats = Reference(ws, min_col=1, min_row=chart_start + 1,
                         max_row=chart_start + len(chart_rows))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.height = 8
        chart.width = 18
        ws.add_chart(chart, f"D{chart_start}")


# ── 3. Per-rule findings with URL + evidence ──────────────────────


def _build_findings(ws: Worksheet, compliance: dict,
                    *, section_filter: set[str] | None = None,
                    title: str = "Findings") -> None:
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:G1")
    ws["A1"] = title
    ws["A1"].font = _TITLE_FONT
    ws.row_dimensions[1].height = 30

    headers = [
        "Section", "Rule", "Severity", "Standard", "URL",
        "Page type", "Evidence",
    ]
    ws.append([])
    ws.append(headers)
    _style_header(ws, 3, len(headers))

    row_idx = 4
    for section in compliance.get("sections", []):
        if section_filter and section.get("key") not in section_filter:
            continue
        for rule in section.get("rules", []):
            if rule.get("count", 0) == 0:
                continue
            refs = rule.get("references", [])
            std = ", ".join(
                f'{r.get("standard","")} {r.get("ref","")}' for r in refs
            )
            for url_row in rule.get("affected_urls", []):
                ws.cell(row=row_idx, column=1, value=section.get("title", ""))
                ws.cell(row=row_idx, column=2, value=rule.get("title", ""))
                sev_cell = ws.cell(row=row_idx, column=3, value=rule.get("severity", ""))
                if rule.get("severity") == "error":
                    sev_cell.font = Font(name="Calibri", size=10, bold=True, color=ERR_RED)
                ws.cell(row=row_idx, column=4, value=std)
                url_cell = ws.cell(row=row_idx, column=5, value=url_row.get("url", ""))
                url_cell.hyperlink = url_row.get("url", "")
                url_cell.font = Font(name="Calibri", size=10, color="0563C1",
                                     underline="single")
                ws.cell(row=row_idx, column=6, value=url_row.get("page_type", ""))
                ws.cell(row=row_idx, column=7, value=url_row.get("evidence", ""))
                for ci in range(1, 8):
                    c = ws.cell(row=row_idx, column=ci)
                    c.alignment = _LEFT
                    c.border = _BORDER
                    if row_idx % 2 == 0:
                        if c.fill.fgColor.rgb in (None, "00000000"):
                            c.fill = _ZEBRA_FILL
                row_idx += 1

    if row_idx == 4:
        ws.cell(row=4, column=1, value="(No violations in this section.)").font = _SUB_FONT

    ws.freeze_panes = "A4"
    if row_idx > 4:
        ws.auto_filter.ref = f"A3:G{row_idx - 1}"
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 60
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 50


# ── Security headers per-URL matrix ───────────────────────────────


def _build_security_matrix(ws: Worksheet, rows: list[dict]) -> None:
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:H1")
    ws["A1"] = "Security Headers Matrix"
    ws["A1"].font = _TITLE_FONT
    ws.row_dimensions[1].height = 30

    headers = [
        "URL", "HSTS", "CSP", "X-Frame-Options",
        "X-Content-Type-Options", "Referrer-Policy",
        "Permissions-Policy", "Mixed-Content?",
    ]
    ws.append([])
    ws.append(headers)
    _style_header(ws, 3, len(headers))

    ok_rows = [r for r in rows if (r.get("status_code") or "") == "200"]
    for i, r in enumerate(ok_rows, start=4):
        ws.cell(row=i, column=1, value=r.get("url", "")).alignment = _LEFT
        ws.cell(row=i, column=1).hyperlink = r.get("url", "")
        ws.cell(row=i, column=1).font = Font(name="Calibri", size=10,
                                             color="0563C1", underline="single")
        vals = [
            r.get("hsts", ""),
            r.get("csp", "")[:60] + ("…" if len(r.get("csp", "")) > 60 else ""),
            r.get("x_frame_options", ""),
            r.get("x_content_type_options", ""),
            r.get("referrer_policy", ""),
            r.get("permissions_policy", "")[:60],
            r.get("has_mixed_content", ""),
        ]
        for ci, v in enumerate(vals, start=2):
            c = ws.cell(row=i, column=ci, value=v)
            c.alignment = _LEFT
            c.border = _BORDER
            if not v or v in ("False", "false"):
                if ci != 8:  # 8 is mixed-content boolean
                    c.fill = _ERR_FILL
                    c.font = Font(name="Calibri", size=10, color=ERR_RED)
            elif v and ci != 8:
                c.fill = _OK_FILL
                c.font = Font(name="Calibri", size=10, color=OK_GREEN)
            if ci == 8 and v in ("True", "true", True):
                c.fill = _ERR_FILL

    ws.freeze_panes = "A4"
    if ok_rows:
        ws.auto_filter.ref = f"A3:H{3 + len(ok_rows)}"
    ws.column_dimensions["A"].width = 60
    for col in "BCDEFGH":
        ws.column_dimensions[col].width = 18


# ── Structured data sheet (JSON-LD types per URL) ─────────────────


def _build_structured_data(ws: Worksheet, rows: list[dict]) -> None:
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:F1")
    ws["A1"] = "Structured Data — JSON-LD Coverage"
    ws["A1"].font = _TITLE_FONT
    ws.row_dimensions[1].height = 30

    headers = [
        "URL", "JSON-LD blocks", "Schema types",
        "Rich-result eligible", "Microdata?", "RDFa?",
    ]
    ws.append([])
    ws.append(headers)
    _style_header(ws, 3, len(headers))

    ok_rows = [r for r in rows if (r.get("status_code") or "") == "200"]
    for i, r in enumerate(ok_rows, start=4):
        url = r.get("url", "")
        types = _row_list(r.get("jsonld_types"))
        rich = _row_list(r.get("jsonld_rich_result_eligible"))
        microdata_n = _safe_int(r.get("microdata_count"))
        rdfa_n = _safe_int(r.get("rdfa_count"))
        url_cell = ws.cell(row=i, column=1, value=url)
        url_cell.hyperlink = url
        url_cell.font = Font(name="Calibri", size=10, color="0563C1",
                             underline="single")
        ws.cell(row=i, column=2, value=_safe_int(r.get("jsonld_count")))
        ws.cell(row=i, column=3, value=", ".join(types))
        ws.cell(row=i, column=4, value=", ".join(rich))
        ws.cell(row=i, column=5, value="yes" if microdata_n > 0 else "")
        ws.cell(row=i, column=6, value="yes" if rdfa_n > 0 else "")
        for ci in range(1, 7):
            c = ws.cell(row=i, column=ci)
            c.alignment = _LEFT
            c.border = _BORDER
            if i % 2 == 0:
                c.fill = _ZEBRA_FILL

    ws.freeze_panes = "A4"
    if ok_rows:
        ws.auto_filter.ref = f"A3:F{3 + len(ok_rows)}"
    ws.column_dimensions["A"].width = 60
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12


# ── Hreflang matrix ──────────────────────────────────────────────


def _build_hreflang_matrix(ws: Worksheet, rows: list[dict]) -> None:
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:F1")
    ws["A1"] = "Hreflang Matrix"
    ws["A1"].font = _TITLE_FONT
    ws.row_dimensions[1].height = 30

    headers = ["URL", "Entry count", "Has x-default?",
               "Self-reference?", "Invalid codes", "Targets"]
    ws.append([])
    ws.append(headers)
    _style_header(ws, 3, len(headers))

    hl_rows = [r for r in rows
               if _safe_int(r.get("hreflang_count")) > 0]
    for i, r in enumerate(hl_rows, start=4):
        url = r.get("url", "")
        entries = _row_list(r.get("hreflang_entries"))
        invalid = _row_list(r.get("hreflang_invalid_codes"))
        url_cell = ws.cell(row=i, column=1, value=url)
        url_cell.hyperlink = url
        url_cell.font = Font(name="Calibri", size=10, color="0563C1",
                             underline="single")
        ws.cell(row=i, column=2, value=_safe_int(r.get("hreflang_count")))
        ws.cell(row=i, column=3, value=str(r.get("hreflang_has_x_default", "")))
        ws.cell(row=i, column=4, value=str(r.get("hreflang_self_reference", "")))
        ws.cell(row=i, column=5, value=", ".join(invalid))
        targets = "; ".join(
            f'{e.get("lang","")}→{e.get("href","")}' for e in entries[:5]
        )
        ws.cell(row=i, column=6, value=targets)
        for ci in range(1, 7):
            c = ws.cell(row=i, column=ci)
            c.alignment = _LEFT
            c.border = _BORDER
            if i % 2 == 0:
                c.fill = _ZEBRA_FILL

    if not hl_rows:
        ws.cell(row=4, column=1,
                value="(No URL on this crawl declared hreflang.)").font = _SUB_FONT

    ws.freeze_panes = "A4"
    if hl_rows:
        ws.auto_filter.ref = f"A3:F{3 + len(hl_rows)}"
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 24
    ws.column_dimensions["F"].width = 80


# ── Technical SEO (canonical + redirect chains) ──────────────────


def _build_technical(ws: Worksheet, rows: list[dict]) -> None:
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:G1")
    ws["A1"] = "Technical SEO — Canonical & Redirect Chains"
    ws["A1"].font = _TITLE_FONT
    ws.row_dimensions[1].height = 30

    headers = [
        "URL", "Status", "Canonical (HTML)", "Canonical (HTTP)",
        "Mismatch?", "Redirect hops", "Redirect chain",
    ]
    ws.append([])
    ws.append(headers)
    _style_header(ws, 3, len(headers))

    for i, r in enumerate(rows, start=4):
        chain = _row_list(r.get("redirect_chain"))
        url_cell = ws.cell(row=i, column=1, value=r.get("url", ""))
        url_cell.hyperlink = r.get("url", "")
        url_cell.font = Font(name="Calibri", size=10, color="0563C1",
                             underline="single")
        ws.cell(row=i, column=2, value=r.get("status_code", ""))
        ws.cell(row=i, column=3, value=r.get("canonical_html", ""))
        ws.cell(row=i, column=4, value=r.get("canonical_http", ""))
        mc = ws.cell(row=i, column=5, value=str(r.get("canonical_mismatch", "")))
        if str(r.get("canonical_mismatch", "")).lower() in ("true", "1"):
            mc.fill = _ERR_FILL
        ws.cell(row=i, column=6, value=_safe_int(r.get("redirect_hops")))
        ws.cell(row=i, column=7, value=" → ".join(chain))
        for ci in range(1, 8):
            c = ws.cell(row=i, column=ci)
            c.alignment = _LEFT
            c.border = _BORDER
            if i % 2 == 0 and c.fill.fgColor.rgb in (None, "00000000"):
                c.fill = _ZEBRA_FILL

    ws.freeze_panes = "A4"
    if rows:
        ws.auto_filter.ref = f"A3:G{3 + len(rows)}"
    ws.column_dimensions["A"].width = 55
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 50
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 80


# ── Content audit ─────────────────────────────────────────────────


def _build_content_audit(ws: Worksheet, rows: list[dict]) -> None:
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:I1")
    ws["A1"] = "Content Audit — Titles, Meta, Readability"
    ws["A1"].font = _TITLE_FONT
    ws.row_dimensions[1].height = 30

    headers = [
        "URL", "Title", "Title pixel width", "Meta description",
        "Meta pixel width", "Word count", "Flesch score",
        "Grade level", "Spelling errors",
    ]
    ws.append([])
    ws.append(headers)
    _style_header(ws, 3, len(headers))

    ok_rows = [r for r in rows if (r.get("status_code") or "") == "200"]
    for i, r in enumerate(ok_rows, start=4):
        url_cell = ws.cell(row=i, column=1, value=r.get("url", ""))
        url_cell.hyperlink = r.get("url", "")
        url_cell.font = Font(name="Calibri", size=10, color="0563C1",
                             underline="single")
        ws.cell(row=i, column=2, value=r.get("title", ""))
        tpx = _safe_int(r.get("title_pixel_width"))
        c_tpx = ws.cell(row=i, column=3, value=tpx)
        if tpx > 580:
            c_tpx.fill = _ERR_FILL
        ws.cell(row=i, column=4, value=r.get("meta_description", ""))
        mpx = _safe_int(r.get("meta_description_pixel_width"))
        c_mpx = ws.cell(row=i, column=5, value=mpx)
        if mpx > 920:
            c_mpx.fill = _ERR_FILL
        ws.cell(row=i, column=6, value=_safe_int(r.get("word_count")))
        try:
            flesch = float(r.get("flesch_score") or 0)
        except (TypeError, ValueError):
            flesch = 0
        c_fl = ws.cell(row=i, column=7, value=round(flesch, 2))
        if 0 < flesch < 30:
            c_fl.fill = _ERR_FILL
        elif 30 <= flesch < 50:
            c_fl.fill = _ZEBRA_FILL
        try:
            grade = float(r.get("grade_level") or 0)
        except (TypeError, ValueError):
            grade = 0
        ws.cell(row=i, column=8, value=round(grade, 2))
        ws.cell(row=i, column=9, value=_safe_int(r.get("spelling_error_count")))
        for ci in range(1, 10):
            c = ws.cell(row=i, column=ci)
            c.alignment = _LEFT
            c.border = _BORDER
            if i % 2 == 0 and c.fill.fgColor.rgb in (None, "00000000"):
                c.fill = _ZEBRA_FILL

    ws.freeze_panes = "A4"
    if ok_rows:
        ws.auto_filter.ref = f"A3:I{3 + len(ok_rows)}"
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 50
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 50
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 12
    ws.column_dimensions["I"].width = 14


# ── Page inventory ────────────────────────────────────────────────


def _build_inventory(ws: Worksheet, rows: list[dict]) -> None:
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:H1")
    ws["A1"] = "Page Inventory — All Crawled URLs"
    ws["A1"].font = _TITLE_FONT
    ws.row_dimensions[1].height = 30

    headers = [
        "URL", "Status", "Page type", "Word count", "Response (ms)",
        "Indexed status", "H1 count", "Images missing alt",
    ]
    ws.append([])
    ws.append(headers)
    _style_header(ws, 3, len(headers))

    for i, r in enumerate(rows, start=4):
        url_cell = ws.cell(row=i, column=1, value=r.get("url", ""))
        url_cell.hyperlink = r.get("url", "")
        url_cell.font = Font(name="Calibri", size=10, color="0563C1",
                             underline="single")
        st_cell = ws.cell(row=i, column=2, value=r.get("status_code", ""))
        if (r.get("status_code") or "") == "200":
            st_cell.font = Font(name="Calibri", size=10, bold=True, color=OK_GREEN)
        else:
            st_cell.font = Font(name="Calibri", size=10, bold=True, color=ERR_RED)
        ws.cell(row=i, column=3, value=r.get("page_type", ""))
        ws.cell(row=i, column=4, value=_safe_int(r.get("word_count")))
        ws.cell(row=i, column=5, value=_safe_int(r.get("response_time_ms")))
        ws.cell(row=i, column=6, value=r.get("indexed_status", ""))
        ws.cell(row=i, column=7, value=_safe_int(r.get("h1_count")))
        ws.cell(row=i, column=8, value=_safe_int(r.get("image_missing_alt")))
        for ci in range(1, 9):
            c = ws.cell(row=i, column=ci)
            c.alignment = _LEFT
            c.border = _BORDER
            if i % 2 == 0 and c.fill.fgColor.rgb in (None, "00000000"):
                c.fill = _ZEBRA_FILL

    ws.freeze_panes = "A4"
    if rows:
        ws.auto_filter.ref = f"A3:H{3 + len(rows)}"
    ws.column_dimensions["A"].width = 55
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 22


# ── Detector catalog ──────────────────────────────────────────────


def _build_catalog(ws: Worksheet, compliance: dict) -> None:
    """Every rule in the catalog with count + why + fix instructions.
    Pulls from compliance payload's sections (already enriched with
    Phase A-D rules)."""
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:F1")
    ws["A1"] = "Detector Catalog — All Rules & Fix Instructions"
    ws["A1"].font = _TITLE_FONT
    ws.row_dimensions[1].height = 30

    headers = ["Rule slug", "Title", "Severity", "Category", "Count", "How to fix"]
    ws.append([])
    ws.append(headers)
    _style_header(ws, 3, len(headers))

    row_idx = 4
    for section in compliance.get("sections", []):
        for rule in section.get("rules", []):
            ws.cell(row=row_idx, column=1, value=rule.get("slug", ""))
            ws.cell(row=row_idx, column=2, value=rule.get("title", ""))
            sev = rule.get("severity", "")
            sev_cell = ws.cell(row=row_idx, column=3, value=sev)
            if sev == "error":
                sev_cell.font = Font(name="Calibri", size=10, bold=True, color=ERR_RED)
            ws.cell(row=row_idx, column=4, value=rule.get("category", ""))
            ws.cell(row=row_idx, column=5, value=rule.get("count", 0))
            ws.cell(row=row_idx, column=6, value=rule.get("how_to_fix", ""))
            for ci in range(1, 7):
                c = ws.cell(row=row_idx, column=ci)
                c.alignment = _LEFT
                c.border = _BORDER
                if row_idx % 2 == 0:
                    c.fill = _ZEBRA_FILL
            row_idx += 1

    ws.freeze_panes = "A4"
    if row_idx > 4:
        ws.auto_filter.ref = f"A3:F{row_idx - 1}"
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 48
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 80


# ── Public entry point ────────────────────────────────────────────


def build_comprehensive_report(
    output_path: Path | None = None,
    *,
    sections: set[str] | None = None,
) -> Path:
    """Build the comprehensive workbook and return its path.

    ``sections`` — set of keys from ALL_SECTIONS to emit. None = emit all.
    """
    from ..audits.runner import _load_rows
    from ..compliance import build_compliance_payload

    rows = _load_rows()
    compliance = build_compliance_payload(max_urls_per_rule=10_000)

    wb = Workbook()
    # Remove default sheet — we add ours in display order.
    if wb.active is not None:
        wb.remove(wb.active)

    emit = sections or set(ALL_SECTIONS)

    if "summary" in emit:
        ws = wb.create_sheet("Executive Summary")
        _build_summary(ws, rows, compliance)
    if "compliance" in emit:
        ws = wb.create_sheet("Compliance Overview")
        _build_compliance_overview(ws, compliance)
    if "wcag" in emit:
        ws = wb.create_sheet("WCAG Findings")
        _build_findings(ws, compliance, section_filter={"wcag"},
                        title="WCAG 2.1 Accessibility Findings")
    if "privacy" in emit:
        ws = wb.create_sheet("Privacy & Cookies")
        _build_findings(ws, compliance, section_filter={"privacy"},
                        title="GDPR / DPDPA — Privacy & Cookies Findings")
    if "security" in emit:
        ws = wb.create_sheet("Security Headers")
        _build_security_matrix(ws, rows)
    if "structured_data" in emit:
        ws = wb.create_sheet("Structured Data")
        _build_structured_data(ws, rows)
    if "hreflang" in emit:
        ws = wb.create_sheet("Hreflang Matrix")
        _build_hreflang_matrix(ws, rows)
    if "technical" in emit:
        ws = wb.create_sheet("Technical SEO")
        _build_technical(ws, rows)
    if "content" in emit:
        ws = wb.create_sheet("Content Audit")
        _build_content_audit(ws, rows)
    if "inventory" in emit:
        ws = wb.create_sheet("Page Inventory")
        _build_inventory(ws, rows)
    if "catalog" in emit:
        ws = wb.create_sheet("Detector Catalog")
        _build_catalog(ws, compliance)

    if output_path is None:
        output_path = settings.reports_path / "crawl_report_comprehensive.xlsx"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path
