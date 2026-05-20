"""Beautified XLSX report bundler.

Produces ``crawl_report.xlsx`` with Bajaj-branded styling. The workbook is
organised by URL category (subdomain + page-type), not just by error type:

  * Summary sheet — KPIs + error pie chart + a category × indexed_status
    pivot block grouping product / knowledge / branch / etc.
  * Per-subdomain overview sheets — www / branch / investmentcorner.
  * Per-category data sheets — only emitted when non-empty.
  * Existing raw sheets — Results / All Errors / 404s / HTTP / Connection /
    Chunked / Console / Discovered — preserved with the five new enrichment
    columns and auto-filter so users can re-slice in Excel.
  * Noise sheet — branch 404s where ``indexed_status != indexed`` — split
    out so the user can ignore it without scrolling past it.
"""
from __future__ import annotations

import csv
import re
from collections import Counter, defaultdict
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from ..conf import settings
from . import url_classifier

# ── Bajaj brand palette ────────────────────────────────────────────────────
BRAND_NAVY = "003DA5"
BRAND_GOLD = "FDB913"
OK_GREEN = "34A853"
OK_GREEN_LT = "E6F4EA"
ERR_RED = "EA4335"
ERR_RED_LT = "FCE8E6"
WARN_ORANGE_LT = "FEF7E0"
ZEBRA = "F8F9FA"
WHITE = "FFFFFF"
TEXT_DARK = "202124"
TEXT_MUTED = "5F6368"

_HEADER_FONT = Font(name="Calibri", size=11, bold=True, color=WHITE)
_SUBHEAD_FONT = Font(name="Calibri", size=10, bold=True, color=TEXT_DARK)
_TITLE_FONT = Font(name="Calibri", size=20, bold=True, color=BRAND_NAVY)
_SUB_FONT = Font(name="Calibri", size=11, color=TEXT_MUTED)
_KPI_LABEL = Font(name="Calibri", size=10, bold=True, color=TEXT_MUTED)
_KPI_VALUE = Font(name="Calibri", size=20, bold=True, color=BRAND_NAVY)

_HEADER_FILL = PatternFill("solid", fgColor=BRAND_NAVY)
_ACCENT_FILL = PatternFill("solid", fgColor=BRAND_GOLD)
_ZEBRA_FILL = PatternFill("solid", fgColor=ZEBRA)
_OK_FILL = PatternFill("solid", fgColor=OK_GREEN_LT)
_ERR_FILL = PatternFill("solid", fgColor=ERR_RED_LT)
_WARN_FILL = PatternFill("solid", fgColor=WARN_ORANGE_LT)  # noqa: F841 reserved for warn rows

_THIN = Side(style="thin", color="E0E4E8")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_CENTER = Alignment(horizontal="center", vertical="center")
_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=False)


def _read_csv(name: str) -> tuple[list[str], list[list[str]]]:
    path = settings.data_path / name
    if not path.exists():
        return [], []
    with open(path, "r", encoding="utf-8") as f:
        reader = csv.reader(f)
        headers = next(reader, [])
        rows = list(reader)
    return headers, rows


def _fit_columns(ws: Worksheet, headers: list[str], rows: list[list[str]],
                 min_w: int = 10, max_w: int = 60) -> None:
    for i, header in enumerate(headers, start=1):
        longest = max(
            [len(header)] + [len(r[i - 1]) if i - 1 < len(r) else 0 for r in rows[:200]]
        )
        ws.column_dimensions[get_column_letter(i)].width = max(min_w, min(max_w, longest + 2))


def _style_header(ws: Worksheet, row: int, cols: int,
                  fill: PatternFill = _HEADER_FILL, font: Font = _HEADER_FONT) -> None:
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = font
        cell.fill = fill
        cell.alignment = _CENTER
        cell.border = _BORDER


def _dump_table(ws: Worksheet, headers: list[str], rows: list[list[str]],
                *, freeze: bool = True, auto_filter: bool = True,
                status_col: str | None = None,
                code_col: str | None = None) -> None:
    """Write a data table with header + zebra body + conditional fills."""
    if not headers:
        ws.cell(row=1, column=1, value="(no data)").font = _SUB_FONT
        return
    status_idx = headers.index(status_col) + 1 if status_col and status_col in headers else None
    code_idx = headers.index(code_col) + 1 if code_col and code_col in headers else None
    idx_idx = headers.index("indexed_status") + 1 if "indexed_status" in headers else None

    ws.append(headers)
    _style_header(ws, 1, len(headers))

    for ri, row in enumerate(rows, start=2):
        padded = (row + [""] * (len(headers) - len(row)))[: len(headers)]
        ws.append(padded)
        base_fill = _ZEBRA_FILL if ri % 2 == 0 else None
        for ci in range(1, len(headers) + 1):
            cell = ws.cell(row=ri, column=ci)
            if base_fill and cell.fill.fgColor.rgb in (None, "00000000"):
                cell.fill = base_fill
            cell.alignment = _LEFT
            cell.border = _BORDER
            cell.font = Font(name="Calibri", size=10, color=TEXT_DARK)

        if status_idx:
            val = padded[status_idx - 1]
            sc = ws.cell(row=ri, column=status_idx)
            if val == "OK":
                sc.fill = _OK_FILL
                sc.font = Font(name="Calibri", size=10, bold=True, color=OK_GREEN)
            elif val and val != "pending":
                sc.fill = _ERR_FILL
                sc.font = Font(name="Calibri", size=10, bold=True, color=ERR_RED)
        if code_idx:
            val = padded[code_idx - 1]
            sc = ws.cell(row=ri, column=code_idx)
            if val == "200":
                sc.font = Font(name="Calibri", size=10, bold=True, color=OK_GREEN)
            elif val.startswith("4") or val.startswith("5"):
                sc.font = Font(name="Calibri", size=10, bold=True, color=ERR_RED)
        if idx_idx:
            val = padded[idx_idx - 1]
            sc = ws.cell(row=ri, column=idx_idx)
            if val == "indexed":
                sc.fill = _OK_FILL
                sc.font = Font(name="Calibri", size=10, bold=True, color=OK_GREEN)
            elif val == "not_indexed":
                sc.fill = _ERR_FILL
                sc.font = Font(name="Calibri", size=10, bold=True, color=ERR_RED)

    if freeze:
        ws.freeze_panes = "A2"
    if auto_filter:
        ws.auto_filter.ref = ws.dimensions
    _fit_columns(ws, headers, rows)
    ws.sheet_view.showGridLines = False


def _build_summary(ws: Worksheet, totals: dict) -> None:
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    for col in ("B", "C", "D", "E", "F"):
        ws.column_dimensions[col].width = 22

    ws["B2"] = "Bajaj Life · Crawl Report"
    ws["B2"].font = _TITLE_FONT
    ws["B3"] = f"Generated {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}  ·  seed: {settings.seed_url}"
    ws["B3"].font = _SUB_FONT
    ws.merge_cells("B2:F2")
    ws.merge_cells("B3:F3")

    for c in range(2, 7):
        ws.cell(row=4, column=c).fill = _ACCENT_FILL
    ws.row_dimensions[4].height = 4

    cards = [
        ("Pages Crawled", totals["pages_crawled"]),
        ("OK (200)", totals["ok_pages"]),
        ("Total Errors", totals["total_errors"]),
        ("404 Errors", totals["errors_404"]),
        ("Discovered Edges", totals["discovered_edges"]),
    ]
    for i, (label, value) in enumerate(cards):
        col = 2 + i
        ws.cell(row=6, column=col, value=label).font = _KPI_LABEL
        ws.cell(row=6, column=col).alignment = _CENTER
        vc = ws.cell(row=7, column=col, value=value)
        vc.font = _KPI_VALUE
        vc.alignment = _CENTER
        for r in (6, 7, 8):
            c = ws.cell(row=r, column=col)
            c.border = Border(left=_THIN, right=_THIN, top=_THIN if r == 6 else None,
                              bottom=_THIN if r == 8 else None)
        ws.row_dimensions[6].height = 22
        ws.row_dimensions[7].height = 32
        ws.row_dimensions[8].height = 4

    ws["B10"] = "Error breakdown"
    ws["B10"].font = _SUBHEAD_FONT
    ws.append([])
    brk_headers = ["Category", "Count"]
    breakdown = [
        ("404 Not Found", totals["errors_404"]),
        ("HTTP Error (non-404)", totals["errors_http"]),
        ("Console Errors (in source)", totals["console_entries"]),
    ]
    start_row = 12
    for i, h in enumerate(brk_headers):
        cell = ws.cell(row=start_row, column=2 + i, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER
        cell.border = _BORDER
    for ri, (cat, count) in enumerate(breakdown, start=start_row + 1):
        ws.cell(row=ri, column=2, value=cat).border = _BORDER
        ws.cell(row=ri, column=3, value=count).border = _BORDER
        ws.cell(row=ri, column=2).font = Font(name="Calibri", size=10, color=TEXT_DARK)
        ws.cell(row=ri, column=3).font = Font(name="Calibri", size=10, bold=True, color=TEXT_DARK)
        if ri % 2 == 0:
            ws.cell(row=ri, column=2).fill = _ZEBRA_FILL
            ws.cell(row=ri, column=3).fill = _ZEBRA_FILL

    if sum(c for _, c in breakdown) > 0:
        chart = PieChart()
        chart.title = "Error distribution"
        labels = Reference(ws, min_col=2, min_row=start_row + 1,
                           max_row=start_row + len(breakdown))
        data = Reference(ws, min_col=3, min_row=start_row,
                         max_row=start_row + len(breakdown))
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(labels)
        chart.height = 9
        chart.width = 14
        chart.dataLabels = DataLabelList(showPercent=True)
        ws.add_chart(chart, "E10")


RAW_SHEETS: list[tuple[str, str, str | None, str | None]] = [
    ("All Results (raw)", "crawl_results.csv", "status", "status_code"),
    ("All 404 Errors (raw)", "crawl_404_errors.csv", None, None),
    ("HTTP Errors", "crawl_errors_httperror.csv", None, None),
    # Connection / Chunked-encoding sheets retired — they were rarely
    # non-empty and not actioned on by any operator workflow.
    ("All Errors", "crawl_errors.csv", None, None),
    ("Console Log", "crawl_console_log.csv", None, None),
    ("Discovered Edges", "crawl_discovered.csv", None, None),
]


def build_report(output_path: Path | None = None) -> Path:
    """Assemble the full XLSX report from current ``data/`` CSVs."""
    out = output_path or (settings.reports_path / "crawl_report.xlsx")
    out.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    summary_ws = wb.active
    summary_ws.title = "Summary"

    res_hdr, res_rows = _read_csv("crawl_results.csv")
    totals = _compute_totals(res_hdr, res_rows)
    breakdown = _compute_breakdown(res_hdr, res_rows)
    _build_summary(summary_ws, totals)
    _build_category_pivot(summary_ws, breakdown)

    # ── Phase 1 audit engine: Health Score + Issues catalogue ──────────
    # Both prepended in front of the Summary tab so the operator opens
    # straight onto the KPI view. The audit runs ONCE here and is passed
    # to both writers; no double-scan.
    try:
        from ..audits import run_all
        from ..services.health_score import compute as compute_health_score

        audit = run_all()
        hs = compute_health_score(audit)
        # Phase 4: per-URL × issue detail sheet (one row per occurrence).
        # Created BEFORE the catalogue sheet so the index parameter
        # pushes it to position 2 after we insert the catalogue at 0.
        detail_ws = wb.create_sheet(title="Issues Detail", index=0)
        _write_issues_detail_sheet(detail_ws, audit)
        issues_ws = wb.create_sheet(title="Issues Catalogue", index=0)
        _write_issues_catalog_sheet(issues_ws, audit)
        health_ws = wb.create_sheet(title="Health Score", index=0)
        _write_health_score_sheet(health_ws, hs)
    except Exception:  # noqa: BLE001 — never let audit failure break the legacy report
        # Audit engine import failure or runtime error must not prevent
        # the rest of the workbook from generating. Logged silently — the
        # operator still gets the Summary + raw sheets they always had.
        pass

    # ── Phase 2: per-competitor sheets + Summary chart ─────────────────
    # One sheet per tracked competitor pulled from the latest gap-
    # pipeline run's GapDeepCrawl rows. Each sheet lists every sampled
    # page with KPI columns the operator can re-slice in Excel.
    # Appended at the end so the legacy sheet order around Summary and
    # per-category is undisturbed.
    try:
        _write_per_competitor_sheets(wb)
    except Exception:  # noqa: BLE001
        pass

    # Add a response-time histogram + status-code bar chart to the
    # existing Summary sheet so the at-a-glance overview gains the
    # charts the operator asked for in Phase 2.
    try:
        _add_summary_charts(summary_ws, res_hdr, res_rows)
    except Exception:  # noqa: BLE001
        pass

    # Per-subdomain overview sheets — readable at-a-glance KPIs per surface.
    for sub in ("www", "branch", "investmentcorner"):
        sub_rows = _filter_rows_by(res_hdr, res_rows, "subdomain", sub)
        if not sub_rows:
            continue
        ws = wb.create_sheet(title=f"{_pretty_sub(sub)} Overview")
        _build_subdomain_overview(ws, sub, res_hdr, sub_rows, breakdown)

    # Per-category data sheets — only emit when there is at least one row.
    for cat in url_classifier.CATEGORY_DEFS:
        cat_rows = _filter_rows_by(res_hdr, res_rows, "category_key", cat["key"])
        if not cat_rows:
            continue
        ws = wb.create_sheet(title=_sheet_name_for_category(cat))
        _dump_table(ws, res_hdr, cat_rows, status_col="status",
                    code_col="status_code")

    # Noise sheet — branch 404s that GSC says are NOT indexed.
    noise_rows = _branch_404_noise_rows(res_hdr, res_rows)
    if noise_rows:
        ws = wb.create_sheet(title="Branch 404 Noise")
        _dump_table(ws, res_hdr, noise_rows, status_col="status",
                    code_col="status_code")

    # Preserved raw sheets for in-Excel re-slicing.
    for sheet_name, csv_file, status_col, code_col in RAW_SHEETS:
        ws = wb.create_sheet(title=sheet_name)
        headers, rows = _read_csv(csv_file)
        _dump_table(ws, headers, rows, status_col=status_col, code_col=code_col)

    wb.save(out)
    return out


# ── New helpers ────────────────────────────────────────────────────────────


def _filter_rows_by(headers: list[str], rows: list[list[str]],
                    column: str, value: str) -> list[list[str]]:
    if column not in headers:
        return []
    idx = headers.index(column)
    return [r for r in rows if idx < len(r) and r[idx] == value]


def _branch_404_noise_rows(headers: list[str],
                           rows: list[list[str]]) -> list[list[str]]:
    if not headers:
        return []
    try:
        sub_i = headers.index("subdomain")
        code_i = headers.index("status_code")
        idx_i = headers.index("indexed_status")
    except ValueError:
        return []
    out = []
    for r in rows:
        if max(sub_i, code_i, idx_i) >= len(r):
            continue
        if r[sub_i] == "branch" and r[code_i] == "404" and r[idx_i] != "indexed":
            out.append(r)
    return out


def _sheet_name_for_category(cat: dict) -> str:
    """Build a stable, Excel-safe sheet name (max 31 chars)."""
    sub = _pretty_sub(cat["subdomain"])
    label = cat["label"]
    raw = f"{sub} · {label}"
    cleaned = raw.replace("·", "-").replace(":", "-").replace("/", "-")
    return cleaned[:31]


def _pretty_sub(sub: str) -> str:
    return {
        "www": "WWW",
        "branch": "Branch",
        "investmentcorner": "InvCorner",
        "external": "External",
    }.get(sub, sub.title())


def _compute_breakdown(headers: list[str],
                       rows: list[list[str]]) -> dict:
    """Per-subdomain × per-category counts for the Summary pivot block."""
    out: dict = {
        "by_subdomain": defaultdict(lambda: Counter()),
        "by_category": defaultdict(lambda: Counter()),
        "noise_404_branch_not_indexed": 0,
    }
    if not headers:
        return out
    try:
        i_sub = headers.index("subdomain")
        i_cat = headers.index("category_key")
        i_code = headers.index("status_code")
        i_idx = headers.index("indexed_status")
        i_src = headers.index("from_sitemap")
    except ValueError:
        return out

    for r in rows:
        if max(i_sub, i_cat, i_code, i_idx, i_src) >= len(r):
            continue
        sub = r[i_sub] or "external"
        cat = r[i_cat] or "unknown"
        code = r[i_code] or ""
        idx = r[i_idx] or "unknown"
        src = r[i_src] or ""

        for bucket_key, bucket in (("by_subdomain", out["by_subdomain"][sub]),
                                   ("by_category",  out["by_category"][cat])):
            bucket["crawled"] += 1
            if code == "200":
                bucket["ok"] += 1
            elif code == "404":
                bucket["errors_404"] += 1
                bucket["errors"] += 1
            elif code and not code.startswith("2"):
                bucket["errors"] += 1
            bucket[f"index_{idx}"] += 1
            if src == "1":
                bucket["from_sitemap"] += 1

        if sub == "branch" and code == "404" and idx != "indexed":
            out["noise_404_branch_not_indexed"] += 1
    return out


def _build_category_pivot(ws: Worksheet, breakdown: dict) -> None:
    """Render the category × indexed pivot block on the Summary sheet."""
    start_row = 20
    ws.cell(row=start_row, column=2, value="Category breakdown").font = _SUBHEAD_FONT

    headers = ["Subdomain", "Category", "Crawled", "OK", "404",
               "Other errors", "Indexed", "Not indexed", "Excluded",
               "Unknown", "From sitemap"]
    header_row = start_row + 1
    for i, h in enumerate(headers, start=2):
        cell = ws.cell(row=header_row, column=i, value=h)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER
        cell.border = _BORDER

    row = header_row + 1
    cat_counts = breakdown["by_category"]
    for cat in url_classifier.CATEGORY_DEFS:
        c = cat_counts.get(cat["key"], Counter())
        if not c:
            continue
        values = [
            _pretty_sub(cat["subdomain"]),
            cat["label"],
            c.get("crawled", 0),
            c.get("ok", 0),
            c.get("errors_404", 0),
            max(c.get("errors", 0) - c.get("errors_404", 0), 0),
            c.get("index_indexed", 0),
            c.get("index_not_indexed", 0),
            c.get("index_excluded", 0),
            c.get("index_unknown", 0),
            c.get("from_sitemap", 0),
        ]
        for i, v in enumerate(values, start=2):
            cell = ws.cell(row=row, column=i, value=v)
            cell.font = Font(name="Calibri", size=10, color=TEXT_DARK)
            cell.border = _BORDER
            if row % 2 == 0:
                cell.fill = _ZEBRA_FILL
        # Conditional red fill on `Not indexed` for product / knowledge
        # categories — these are the ones we genuinely care about.
        ni_cell = ws.cell(row=row, column=8)
        if cat["key"].startswith("product") or cat["key"] == "knowledge":
            if (ni_cell.value or 0) > 0:
                ni_cell.fill = _ERR_FILL
                ni_cell.font = Font(name="Calibri", size=10, bold=True, color=ERR_RED)
        row += 1

    ws.cell(row=row + 1, column=2,
            value=f"Noise: branch 404s NOT indexed = "
                  f"{breakdown.get('noise_404_branch_not_indexed', 0)}"
            ).font = _SUB_FONT


def _build_subdomain_overview(ws: Worksheet, sub: str,
                              headers: list[str], rows: list[list[str]],
                              breakdown: dict) -> None:
    """Per-subdomain summary + the top 20 problem URLs on that surface."""
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 2
    for col in ("B", "C", "D", "E", "F"):
        ws.column_dimensions[col].width = 22

    title = f"{_pretty_sub(sub)} surface"
    ws["B2"] = title
    ws["B2"].font = _TITLE_FONT
    ws["B3"] = f"Generated {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}  ·  {len(rows)} rows"
    ws["B3"].font = _SUB_FONT
    ws.merge_cells("B2:F2")
    ws.merge_cells("B3:F3")

    counts = breakdown["by_subdomain"].get(sub, Counter())
    cards = [
        ("Pages", counts.get("crawled", len(rows))),
        ("OK (200)", counts.get("ok", 0)),
        ("Errors", counts.get("errors", 0)),
        ("Indexed", counts.get("index_indexed", 0)),
        ("Not Indexed", counts.get("index_not_indexed", 0)),
    ]
    for i, (label, value) in enumerate(cards):
        col = 2 + i
        ws.cell(row=6, column=col, value=label).font = _KPI_LABEL
        ws.cell(row=6, column=col).alignment = _CENTER
        vc = ws.cell(row=7, column=col, value=value)
        vc.font = _KPI_VALUE
        vc.alignment = _CENTER

    # Top problem URLs — anything non-200 on this surface.
    code_i = headers.index("status_code") if "status_code" in headers else None
    url_i = headers.index("url") if "url" in headers else 0
    idx_i = headers.index("indexed_status") if "indexed_status" in headers else None
    cat_i = headers.index("category_key") if "category_key" in headers else None
    problems = []
    if code_i is not None:
        for r in rows:
            if code_i < len(r) and r[code_i] not in ("200", ""):
                problems.append([
                    r[url_i] if url_i < len(r) else "",
                    r[code_i],
                    r[idx_i] if idx_i is not None and idx_i < len(r) else "",
                    r[cat_i] if cat_i is not None and cat_i < len(r) else "",
                ])
    problems = problems[:20]

    if problems:
        block = 11
        for i, h in enumerate(["URL", "Status", "Indexed", "Category"], start=2):
            cell = ws.cell(row=block, column=i, value=h)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = _CENTER
            cell.border = _BORDER
        for ri, row in enumerate(problems, start=block + 1):
            for ci, v in enumerate(row, start=2):
                c = ws.cell(row=ri, column=ci, value=v)
                c.font = Font(name="Calibri", size=10, color=TEXT_DARK)
                c.border = _BORDER
                if ri % 2 == 0:
                    c.fill = _ZEBRA_FILL


def _compute_totals(res_hdr: list[str], res_rows: list[list[str]]) -> dict:
    def count(name: str) -> int:
        _, rows = _read_csv(name)
        return len(rows)

    ok = 0
    if res_hdr:
        try:
            idx = res_hdr.index("status_code")
            ok = sum(1 for r in res_rows if idx < len(r) and r[idx] == "200")
        except ValueError:
            pass
    return {
        "pages_crawled": len(res_rows),
        "ok_pages": ok,
        "total_errors": count("crawl_errors.csv"),
        "errors_404": count("crawl_404_errors.csv"),
        "errors_http": count("crawl_errors_httperror.csv"),
        "console_entries": count("crawl_console_log.csv"),
        "discovered_edges": count("crawl_discovered.csv"),
    }


# ── Phase 1: Health Score + Issues Catalogue sheets ────────────────────
#
# Both sheets are prepended to the workbook in build_report so they appear
# before the existing Summary tab. Reuses the Bajaj brand palette + cell
# styling helpers defined at the top of this module — no new styling
# constants required.

_TIER_FILL = {
    "Excellent": PatternFill("solid", fgColor=OK_GREEN),
    "Good":      PatternFill("solid", fgColor=BRAND_NAVY),
    "Fair":      PatternFill("solid", fgColor=BRAND_GOLD),
    "Weak":      PatternFill("solid", fgColor=ERR_RED),
}


def _write_health_score_sheet(ws: Worksheet, hs) -> None:
    """Top-of-workbook KPI sheet.

    Mirrors the Health Score widget rendered in the dashboard:
      * Big score (0-100) + tier badge.
      * Severity counts (errors / warnings / notices) + distinct-type counts.
      * Top-5 most-affecting error issues.
      * Category coverage tile grid.
      * Formula footnote.

    ``hs`` is a ``services.health_score.HealthScore`` instance. The import
    path lives inside ``build_report`` so audit-engine failure can't
    prevent the legacy report from generating.
    """
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 28
    ws.column_dimensions["F"].width = 14

    ws.merge_cells("A1:F1")
    ws["A1"] = "Health Score"
    ws["A1"].font = _TITLE_FONT
    ws["A1"].alignment = _LEFT

    ws.merge_cells("A2:F2")
    finished = hs.finished_at[:19] if hs.finished_at else ""
    ws["A2"] = (
        f"Computed {finished}  -  "
        f"{hs.urls_without_error:,} of {hs.total_urls:,} URLs without errors"
    )
    ws["A2"].font = _SUB_FONT
    ws["A2"].alignment = _LEFT

    ws.merge_cells("A4:B6")
    score_cell = ws["A4"]
    score_cell.value = hs.score
    score_cell.font = Font(name="Calibri", size=60, bold=True, color=BRAND_NAVY)
    score_cell.alignment = _CENTER

    ws.merge_cells("A7:B7")
    tier_cell = ws["A7"]
    tier_cell.value = hs.tier.upper()
    tier_cell.font = Font(name="Calibri", size=12, bold=True, color=WHITE)
    tier_cell.alignment = _CENTER
    tier_cell.fill = _TIER_FILL.get(hs.tier, _HEADER_FILL)

    sev_labels = (
        ("D4", "ERRORS", hs.severity_counts.get("error", 0),
         hs.issue_type_counts.get("error", 0), ERR_RED),
        ("D6", "WARNINGS", hs.severity_counts.get("warning", 0),
         hs.issue_type_counts.get("warning", 0), "F59E0B"),
        ("D8", "NOTICES", hs.severity_counts.get("notice", 0),
         hs.issue_type_counts.get("notice", 0), BRAND_NAVY),
    )
    for cell, label, count, types, color in sev_labels:
        ws[cell] = label
        ws[cell].font = Font(name="Calibri", size=9, bold=True, color=TEXT_MUTED)
        right_cell = cell.replace("D", "E")
        plural = "s" if types != 1 else ""
        ws[right_cell] = f"{count:,}   ({types} type{plural})"
        ws[right_cell].font = Font(name="Calibri", size=14, bold=True, color=color)

    row = 11
    ws.cell(row=row, column=1, value="TOP ERRORS BY AFFECTED URLs").font = (
        Font(name="Calibri", size=10, bold=True, color=TEXT_MUTED)
    )
    row += 1
    for i, h in enumerate(("Issue", "Severity", "Category", "Affected URLs"), start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL
        c.alignment = _CENTER
        c.border = _BORDER
    row += 1
    for issue in hs.top_errors:
        ws.cell(row=row, column=1, value=issue["title"]).font = Font(
            name="Calibri", size=11, color=TEXT_DARK,
        )
        sev = ws.cell(row=row, column=2, value=issue["severity"].title())
        sev.fill = _ERR_FILL
        sev.alignment = _CENTER
        ws.cell(row=row, column=3, value=issue["category"]).alignment = _CENTER
        cnt = ws.cell(row=row, column=4, value=issue["count"])
        cnt.alignment = _CENTER
        cnt.font = Font(name="Calibri", size=11, bold=True, color=ERR_RED)
        row += 1

    row += 1
    ws.cell(row=row, column=1, value="ISSUE TYPES PER CATEGORY").font = (
        Font(name="Calibri", size=10, bold=True, color=TEXT_MUTED)
    )
    row += 1
    cats = sorted(hs.category_counts.items(), key=lambda kv: -kv[1])
    for i, (cat, n) in enumerate(cats):
        col = (i % 3) * 2 + 1
        r = row + (i // 3)
        label = ws.cell(row=r, column=col, value=cat.upper())
        label.font = Font(name="Calibri", size=9, bold=True, color=TEXT_MUTED)
        plural = "s" if n != 1 else ""
        val = ws.cell(row=r, column=col + 1, value=f"{n} type{plural}")
        val.font = Font(name="Calibri", size=12, bold=True, color=BRAND_NAVY)

    final = row + (len(cats) // 3) + 3
    ws.merge_cells(start_row=final, start_column=1, end_row=final, end_column=6)
    foot = ws.cell(row=final, column=1, value=f"Formula:  {hs.formula}")
    foot.font = Font(name="Calibri", size=9, italic=True, color=TEXT_MUTED)


def _write_issues_catalog_sheet(ws: Worksheet, audit) -> None:
    """Issue triage inbox in spreadsheet form.

    One row per distinct issue type that fired. Columns:
    Slug / Severity / Category / Issue / Affected URLs / Why / How to fix.

    Errors sort to the top; within severity, by URL count descending.
    Severity column conditionally-filled (red/orange/blue) so the eye
    can land on errors immediately.

    ``audit`` is an ``audits.runner.AuditResult``.
    """
    ws.sheet_view.showGridLines = False

    headers = ("Slug", "Severity", "Category", "Issue", "Affected URLs",
               "Why it matters", "How to fix")
    widths = (24, 12, 18, 38, 16, 60, 60)
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.merge_cells("A1:G1")
    ws["A1"] = "Issues Catalogue"
    ws["A1"].font = _TITLE_FONT
    ws["A1"].alignment = _LEFT

    ws.merge_cells("A2:G2")
    finished = audit.finished_at[:19] if audit.finished_at else ""
    ws["A2"] = (
        f"Computed {finished}  -  "
        f"{audit.total_urls:,} URLs scanned  -  "
        f"{audit.urls_with_any_error:,} URLs with at least one error"
    )
    ws["A2"].font = _SUB_FONT
    ws["A2"].alignment = _LEFT

    severity_order = {"error": 0, "warning": 1, "notice": 2}
    occs = sorted(
        [o for o in audit.occurrences if o.count > 0],
        key=lambda o: (severity_order[o.issue.severity], -o.count),
    )

    header_row = 4
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=i, value=h)
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL
        c.alignment = _CENTER
        c.border = _BORDER

    sev_fill_map = {
        "error": _ERR_FILL,
        "warning": _WARN_FILL,
        "notice": PatternFill("solid", fgColor="E6EEF9"),
    }
    sev_text_color = {
        "error": ERR_RED,
        "warning": "F59E0B",
        "notice": BRAND_NAVY,
    }

    row = header_row + 1
    for occ in occs:
        issue = occ.issue
        zebra = _ZEBRA_FILL if (row - header_row) % 2 == 0 else None
        cells = [
            (1, issue.slug,
             Font(name="Consolas", size=10, color=TEXT_DARK)),
            (2, issue.severity.title(),
             Font(name="Calibri", size=10, bold=True,
                  color=sev_text_color[issue.severity])),
            (3, issue.category,
             Font(name="Calibri", size=10, color=TEXT_MUTED)),
            (4, issue.title,
             Font(name="Calibri", size=11, bold=True, color=TEXT_DARK)),
            (5, occ.count,
             Font(name="Calibri", size=11, bold=True, color=BRAND_NAVY)),
            (6, issue.why,
             Font(name="Calibri", size=10, color=TEXT_DARK)),
            (7, issue.how_to_fix,
             Font(name="Calibri", size=10, color=TEXT_DARK)),
        ]
        for col, val, font in cells:
            c = ws.cell(row=row, column=col, value=val)
            c.font = font
            c.alignment = Alignment(
                horizontal="center" if col in (2, 3, 5) else "left",
                vertical="top",
                wrap_text=col in (4, 6, 7),
            )
            c.border = _BORDER
            if col == 2:
                c.fill = sev_fill_map[issue.severity]
            elif zebra is not None:
                c.fill = zebra
        ws.row_dimensions[row].height = 64
        row += 1

    ws.freeze_panes = ws.cell(row=header_row + 1, column=5)
    ws.auto_filter.ref = (
        f"A{header_row}:G{row - 1}"
        if row > header_row + 1
        else f"A{header_row}:G{header_row}"
    )


# ── Phase 2: per-competitor sheets + Summary charts ────────────────────


def _safe_sheet_title(raw: str) -> str:
    """Excel sheet names cap at 31 chars and ban these characters:
    ``: \\ / ? * [ ]``. Sanitises the competitor domain into a
    spreadsheet-valid title."""
    cleaned = re.sub(r"[:\\/?*\[\]]", "-", raw or "")
    cleaned = cleaned[:31].rstrip()
    return cleaned or "Competitor"


def _write_per_competitor_sheets(wb) -> None:
    """One sheet per top-10 competitor from the latest gap-pipeline run.

    Each row = one URL the competitor crawler sampled from that domain.
    Columns mirror the per-competitor Page Explorer view: URL, title,
    meta_description, page_type, word_count, schema, internal/external
    link counts, response time, PageSpeed, LCP, CLS, INP.

    Skipped silently when:
      - apps.seo_ai is unavailable (e.g., migrations not applied)
      - no GapDeepCrawl rows for the latest run
      - a competitor row has an empty profile (crawl failed for that
        domain — surface only the error column to flag it)
    """
    from apps.seo_ai.models import GapDeepCrawl, GapPipelineRun

    run = GapPipelineRun.objects.order_by("-started_at").first()
    if run is None:
        return
    crawls = list(
        GapDeepCrawl.objects.filter(run=run).order_by("is_us", "domain")
    )
    if not crawls:
        return

    columns = (
        ("url", 60),
        ("title", 50),
        ("meta_description", 60),
        ("page_type", 14),
        ("word_count", 12),
        ("has_schema", 12),
        ("schema_types", 30),
        ("response_time_ms", 16),
        ("internal_link_count", 18),
        ("external_link_count", 18),
        ("pagespeed_score", 16),
        ("lcp_ms", 12),
        ("cls", 10),
        ("inp_ms", 12),
        ("last_modified", 22),
    )

    for c in crawls:
        title_prefix = "Us - " if c.is_us else "Comp - "
        sheet_title = _safe_sheet_title(title_prefix + c.domain)
        ws = wb.create_sheet(title=sheet_title)
        ws.sheet_view.showGridLines = False

        # Header / context block
        ws.merge_cells(f"A1:{get_column_letter(len(columns))}1")
        ws["A1"] = c.domain + (" (us)" if c.is_us else "")
        ws["A1"].font = _TITLE_FONT
        ws["A1"].alignment = _LEFT

        ws.merge_cells(f"A2:{get_column_letter(len(columns))}2")
        ws["A2"] = (
            f"Sampled {c.pages_attempted} pages ({c.pages_ok} OK) "
            f"from a sitemap of {c.sitemap_url_count:,} URLs"
        )
        ws["A2"].font = _SUB_FONT

        if c.error:
            ws.merge_cells(f"A4:{get_column_letter(len(columns))}4")
            err_cell = ws["A4"]
            err_cell.value = f"Crawl error: {c.error}"
            err_cell.font = Font(name="Calibri", size=11, color=ERR_RED)
            err_cell.fill = _ERR_FILL
            continue

        samples = (c.profile or {}).get("sample_pages") or []
        if not samples:
            ws["A4"] = "No sample pages captured."
            ws["A4"].font = _SUB_FONT
            continue

        # Column widths + header row
        for i, (name, width) in enumerate(columns, start=1):
            ws.column_dimensions[get_column_letter(i)].width = width
        for i, (name, _) in enumerate(columns, start=1):
            cell = ws.cell(row=4, column=i, value=name)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = _CENTER
            cell.border = _BORDER

        # Data rows
        for ridx, sample in enumerate(samples, start=5):
            zebra = _ZEBRA_FILL if (ridx - 5) % 2 == 1 else None
            for cidx, (name, _) in enumerate(columns, start=1):
                value = sample.get(name)
                if name == "schema_types" and isinstance(value, list):
                    value = ", ".join(str(v) for v in value[:10])
                elif name == "has_schema":
                    value = "yes" if value else "no"
                if value is None:
                    value = ""
                cell = ws.cell(row=ridx, column=cidx, value=value)
                cell.alignment = Alignment(
                    horizontal=("left" if cidx in (1, 2, 3, 7) else "center"),
                    vertical="center",
                    wrap_text=cidx in (2, 3),
                )
                cell.border = _BORDER
                if zebra is not None:
                    cell.fill = zebra
            ws.row_dimensions[ridx].height = 24

        # Freeze headers, add auto-filter
        ws.freeze_panes = ws.cell(row=5, column=2)
        ws.auto_filter.ref = (
            f"A4:{get_column_letter(len(columns))}{4 + len(samples)}"
        )


def _add_summary_charts(ws, res_hdr: list[str], res_rows: list[list[str]]) -> None:
    """Append two charts to the existing Summary sheet:

      * Status-code distribution (bar chart) — shows the OK / 4xx /
        5xx / 0 split visually.
      * Response-time histogram (bar chart) — buckets pages into 5
        speed tiers (<200ms, 200-500ms, 500ms-1s, 1-3s, >3s).

    Both charts placed below the existing category pivot so the
    layout above is unchanged.
    """
    if not res_rows:
        return
    from collections import Counter
    from openpyxl.chart import BarChart, Reference

    status_idx = res_hdr.index("status_code") if "status_code" in res_hdr else -1
    rt_idx = (
        res_hdr.index("response_time_ms")
        if "response_time_ms" in res_hdr
        else -1
    )

    # Find an empty row below the existing pivot. Iterate up to row 50
    # safely — the pivot block never approaches that depth.
    start_row = 32
    while ws.cell(row=start_row, column=1).value is not None and start_row < 80:
        start_row += 1

    if status_idx >= 0:
        counts = Counter()
        for r in res_rows:
            if status_idx < len(r):
                code = (r[status_idx] or "").strip() or "unknown"
                counts[code] += 1

        ws.cell(row=start_row, column=1, value="STATUS CODE DISTRIBUTION").font = (
            Font(name="Calibri", size=10, bold=True, color=TEXT_MUTED)
        )
        header_row = start_row + 1
        ws.cell(row=header_row, column=1, value="Status")
        ws.cell(row=header_row, column=2, value="Count")
        for c in (1, 2):
            cell = ws.cell(row=header_row, column=c)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = _CENTER
            cell.border = _BORDER

        ordered = sorted(counts.items(), key=lambda kv: -kv[1])
        for i, (code, n) in enumerate(ordered, start=1):
            ws.cell(row=header_row + i, column=1, value=code).alignment = _CENTER
            ws.cell(row=header_row + i, column=2, value=n).alignment = _CENTER

        chart = BarChart()
        chart.type = "col"
        chart.style = 11
        chart.title = "Status code distribution"
        chart.y_axis.title = "URLs"
        chart.x_axis.title = "HTTP status"
        data = Reference(
            ws,
            min_col=2, min_row=header_row,
            max_row=header_row + len(ordered), max_col=2,
        )
        cats = Reference(
            ws,
            min_col=1, min_row=header_row + 1,
            max_row=header_row + len(ordered),
        )
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.width = 16
        chart.height = 8
        ws.add_chart(chart, ws.cell(row=start_row, column=4).coordinate)

        start_row = header_row + len(ordered) + 3

    if rt_idx >= 0:
        buckets = [
            ("<200 ms",   lambda v: v < 200),
            ("200-500 ms", lambda v: 200 <= v < 500),
            ("500ms-1s",  lambda v: 500 <= v < 1000),
            ("1-3 s",     lambda v: 1000 <= v < 3000),
            (">3 s",      lambda v: v >= 3000),
        ]
        counts = [0] * len(buckets)
        for r in res_rows:
            if rt_idx >= len(r):
                continue
            try:
                v = int(r[rt_idx] or 0)
            except ValueError:
                continue
            if v <= 0:
                continue
            for i, (_, pred) in enumerate(buckets):
                if pred(v):
                    counts[i] += 1
                    break

        ws.cell(row=start_row, column=1, value="RESPONSE TIME HISTOGRAM").font = (
            Font(name="Calibri", size=10, bold=True, color=TEXT_MUTED)
        )
        header_row = start_row + 1
        ws.cell(row=header_row, column=1, value="Bucket")
        ws.cell(row=header_row, column=2, value="Count")
        for c in (1, 2):
            cell = ws.cell(row=header_row, column=c)
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
            cell.alignment = _CENTER
            cell.border = _BORDER
        for i, ((label, _), n) in enumerate(zip(buckets, counts), start=1):
            ws.cell(row=header_row + i, column=1, value=label).alignment = _CENTER
            ws.cell(row=header_row + i, column=2, value=n).alignment = _CENTER

        chart = BarChart()
        chart.type = "col"
        chart.style = 10
        chart.title = "Response time distribution"
        chart.y_axis.title = "URLs"
        chart.x_axis.title = "Response time"
        data = Reference(
            ws,
            min_col=2, min_row=header_row,
            max_row=header_row + len(buckets), max_col=2,
        )
        cats = Reference(
            ws,
            min_col=1, min_row=header_row + 1,
            max_row=header_row + len(buckets),
        )
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.width = 16
        chart.height = 8
        ws.add_chart(chart, ws.cell(row=start_row, column=4).coordinate)


# ── Phase 4: per-URL × issue detail sheet ─────────────────────────────


def _write_issues_detail_sheet(ws: Worksheet, audit) -> None:
    """One row per (issue × affected URL) — the triage drill-down.

    Different shape from the Phase 1 "Issues Catalogue" sheet which
    has one row per issue type. The Detail sheet lets the operator
    sort/filter the full URL list in Excel: all errors on the branch
    subdomain, all warnings touching the /term-insurance/ path, etc.

    Affected URLs are already capped at 1000 per issue in the audit
    runner; with ~30-50 active issue types, the detail sheet stays
    under ~50,000 rows — well within Excel's 1M limit.
    """
    ws.sheet_view.showGridLines = False
    headers = (
        "Issue", "Severity", "Category", "URL", "Title",
        "Status", "Page type", "Subdomain", "Words", "Response (ms)",
    )
    widths = (38, 12, 16, 60, 40, 10, 14, 14, 10, 14)
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.merge_cells(f"A1:{get_column_letter(len(headers))}1")
    ws["A1"] = "Issues Detail (per URL)"
    ws["A1"].font = _TITLE_FONT
    ws["A1"].alignment = _LEFT

    ws.merge_cells(f"A2:{get_column_letter(len(headers))}2")
    finished = audit.finished_at[:19] if audit.finished_at else ""
    severity_breakdown = audit.severity_counts()
    ws["A2"] = (
        f"Computed {finished}  -  "
        f"{audit.total_urls:,} URLs scanned  -  "
        f"{severity_breakdown['error']:,} errors + "
        f"{severity_breakdown['warning']:,} warnings + "
        f"{severity_breakdown['notice']:,} notices"
    )
    ws["A2"].font = _SUB_FONT
    ws["A2"].alignment = _LEFT

    header_row = 4
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=i, value=h)
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL
        c.alignment = _CENTER
        c.border = _BORDER

    sev_fill = {
        "error": _ERR_FILL,
        "warning": _WARN_FILL,
        "notice": PatternFill("solid", fgColor="E6EEF9"),
    }
    sev_text = {
        "error": ERR_RED,
        "warning": "F59E0B",
        "notice": BRAND_NAVY,
    }
    severity_order = {"error": 0, "warning": 1, "notice": 2}

    # Sort issue occurrences errors-first, then by count desc, then
    # iterate affected URLs within each occurrence.
    occs = sorted(
        [o for o in audit.occurrences if o.count > 0],
        key=lambda o: (severity_order[o.issue.severity], -o.count),
    )

    row = header_row + 1
    for occ in occs:
        issue = occ.issue
        for affected in occ.affected_urls:
            url = (affected.get("url") or "").strip()
            if not url:
                continue
            zebra = _ZEBRA_FILL if (row - header_row) % 2 == 0 else None
            values = [
                issue.title,
                issue.severity.title(),
                issue.category,
                url,
                (affected.get("title") or "").strip(),
                (affected.get("status_code") or "").strip(),
                (affected.get("page_type") or "").strip(),
                (affected.get("subdomain") or "").strip(),
                affected.get("word_count") or "",
                affected.get("response_time_ms") or "",
            ]
            for col, val in enumerate(values, start=1):
                c = ws.cell(row=row, column=col, value=val)
                c.font = Font(
                    name="Consolas" if col == 4 else "Calibri",
                    size=10,
                    bold=(col == 2),
                    color=sev_text[issue.severity] if col == 2 else TEXT_DARK,
                )
                c.alignment = Alignment(
                    horizontal="center" if col in (2, 3, 6, 7, 8, 9, 10) else "left",
                    vertical="top",
                    wrap_text=col in (1, 5),
                )
                c.border = _BORDER
                if col == 2:
                    c.fill = sev_fill[issue.severity]
                elif zebra is not None:
                    c.fill = zebra
            row += 1

    if row > header_row + 1:
        ws.freeze_panes = ws.cell(row=header_row + 1, column=5)
        ws.auto_filter.ref = (
            f"A{header_row}:{get_column_letter(len(headers))}{row - 1}"
        )

