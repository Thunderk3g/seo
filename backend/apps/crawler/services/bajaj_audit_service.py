import os
import re
import time
import requests
import logging
import threading
import concurrent.futures
from datetime import datetime
from collections import deque
from urllib.parse import urljoin, urlparse

import pandas as pd
from bs4 import BeautifulSoup
from django.conf import settings
import urllib3

# Suppress insecure request warnings from urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# ── LOGGING ────────────────────────────────────────────────────────────────────
log = logging.getLogger("seo")

# ── CONFIG ─────────────────────────────────────────────────────────────────────
# Default config, can be overridden by the command arguments
DEFAULT_INPUT_FILE = r'C:\Users\Diwakar.Adhikari01\Desktop\SEO\Partners_Site_List.xlsx'
DEFAULT_OUTPUT_EXCEL = 'Bajaj_Allianz_Life_Deep_Audit.xlsx'
DEFAULT_PROGRESS_CSV = '_crawl_progress.csv'

# Match criteria
TEXT_MATCH = 'bajaj allianz life'          # in visible page text
HREF_MATCH = 'bajajallianzlife.com'        # in any <a href>

# Crawl settings Default
MAX_DEPTH = 3       # levels deep from homepage
MAX_PAGES_PER_SITE = 500     # safety cap per domain
MAX_SITEMAP_URLS = 1000    # max URLs to pull from sitemap
REQUEST_TIMEOUT = 12
DELAY_BETWEEN_REQ = 0.3     # seconds between requests (polite)
DELAY_BETWEEN_DOMAINS = 1.0  # seconds between domains

HEADERS = {
    'User-Agent': (
        'Mozilla/5.0 (Windows NT 10.0; Win64; x64) '
        'AppleWebKit/537.36 (KHTML, like Gecko) '
        'Chrome/121.0.0.0 Safari/537.36'
    ),
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
    'Accept-Language': 'en-US,en;q=0.5',
    'Accept-Encoding': 'gzip, deflate',
    'Connection': 'keep-alive',
}


class BajajAuditService:
    """Service to crawl partner websites for Bajaj Allianz Life references.
    
    This encapsulates the procedural deep crawler script into a reusable Django service.
    """

    def __init__(self, input_file=DEFAULT_INPUT_FILE, output_file=DEFAULT_OUTPUT_EXCEL, progress_csv=DEFAULT_PROGRESS_CSV):
        self.input_file = input_file
        self.output_excel = output_file
        self.progress_csv = progress_csv

    # ── HELPERS ────────────────────────────────────────────────────────────────────
    
    @staticmethod
    def normalize_domain(raw):
        """Strip protocol/path, return bare domain."""
        raw = str(raw).strip().lower()
        if not raw.startswith('http'):
            raw = 'https://' + raw
        try:
            return urlparse(raw).netloc.replace('www.', '')
        except:
            return raw

    @staticmethod
    def base_url(domain):
        return f'https://{domain}'

    @staticmethod
    def same_domain(url, domain):
        """True if URL belongs to the same domain (or www. variant)."""
        try:
            host = urlparse(url).netloc.lower().replace('www.', '')
            return host == domain.lower().replace('www.', '')
        except:
            return False

    def fetch(self, url, timeout=REQUEST_TIMEOUT):
        """Fetch a URL, return (html_text, final_url, status_code) or (None, url, 0)."""
        session = requests.Session()
        adapter = requests.adapters.HTTPAdapter(max_retries=1)
        session.mount('http://', adapter)
        session.mount('https://', adapter)
        
        # Enforce both connection and read timeouts strictly
        req_timeout = (timeout, timeout)
        
        try:
            # Drop stream=True to prevent iter_content from hanging in _sslObj.read indefinitely
            r = session.get(url, headers=HEADERS, timeout=req_timeout,
                             allow_redirects=True, verify=False, stream=False)
                             
            # Fast fail if it's explicitly not HTML (e.g. PDF, Images)
            content_type = r.headers.get('Content-Type', '').lower()
            allowed = ['text/html', 'application/xhtml+xml', 'text/xml', 'application/xml']
            if not any(a in content_type for a in allowed) and content_type:
                log.info(f"Skipping non-HTML/XML content: {url} ({content_type})")
                r.close()
                session.close()
                return None, url, 0

            # Only process up to 4MB of text
            raw = r.content[:4_000_000]
            r.close()
            session.close()
            return raw.decode('utf-8', errors='replace'), r.url, r.status_code
            
        except requests.exceptions.Timeout:
            return None, url, -1
        except requests.exceptions.ConnectionError:
            return None, url, -2
        except requests.exceptions.SSLError:
            # Retry with http
            try:
                http_url = url.replace('https://', 'http://', 1)
                r = session.get(http_url, headers=HEADERS, timeout=req_timeout,
                                 allow_redirects=True, verify=False, stream=False)
                                 
                content_type = r.headers.get('Content-Type', '').lower()
                allowed = ['text/html', 'application/xhtml+xml', 'text/xml', 'application/xml']
                if not any(a in content_type for a in allowed) and content_type:
                    r.close()
                    session.close()
                    return None, url, 0

                raw = r.content[:4_000_000]
                r.close()
                session.close()
                return raw.decode('utf-8', errors='replace'), r.url, r.status_code
            except Exception as e:
                log.warning(f"Fetch failed on HTTP retry for {url}: {e}")
                session.close()
                return None, url, -3
        except Exception as e:
            log.warning(f"Fetch failed for {url}: {e}")
            session.close()
            return None, url, -4

    # ── SITEMAP DISCOVERY ──────────────────────────────────────────────────────────

    def get_sitemap_urls(self, domain):
        """
        Try common sitemap locations.
        Returns list of page URLs found, or empty list if no sitemap.
        """
        candidates = [
            f'https://{domain}/sitemap.xml',
            f'https://{domain}/sitemap_index.xml',
            f'https://{domain}/sitemap/',
            f'https://{domain}/sitemap.php',
            f'https://{domain}/sitemap.txt',
            f'http://{domain}/sitemap.xml',
        ]
        all_urls = []

        for sm_url in candidates:
            html, final_url, code = self.fetch(sm_url)
            if not html or code not in (200,):
                continue

            # Sitemap index — contains <sitemap> tags
            if '<sitemapindex' in html.lower():
                try:
                    soup = BeautifulSoup(html, 'lxml-xml')
                except Exception as e:
                    log.warning(f"Sitemap parsing failed for {sm_url}: {e}")
                    continue
                sub_locs = [t.text.strip() for t in soup.find_all('loc')]
                for sub in sub_locs[:10]:   # fetch up to 10 sub-sitemaps
                    sub_html, _, sub_code = self.fetch(sub)
                    if sub_html and sub_code == 200:
                        try:
                            sub_soup = BeautifulSoup(sub_html, 'lxml-xml')
                        except Exception as e:
                            log.warning(f"Sub-sitemap parsing failed for {sub}: {e}")
                            continue
                        for loc in sub_soup.find_all('loc'):
                            u = loc.text.strip()
                            if self.same_domain(u, domain) and u not in all_urls:
                                all_urls.append(u)
                            if len(all_urls) >= MAX_SITEMAP_URLS:
                                break
                    if len(all_urls) >= MAX_SITEMAP_URLS:
                        break
                break

            # Regular sitemap — contains <url><loc> tags
            elif '<urlset' in html.lower() or '<loc>' in html.lower():
                try:
                    soup = BeautifulSoup(html, 'lxml-xml')
                except Exception as e:
                    log.warning(f"Sitemap parsing failed for {sm_url}: {e}")
                    continue
                for loc in soup.find_all('loc'):
                    u = loc.text.strip()
                    if self.same_domain(u, domain) and u not in all_urls:
                        all_urls.append(u)
                    if len(all_urls) >= MAX_SITEMAP_URLS:
                        break
                break

            # Plain text sitemap
            elif html.strip().startswith('http'):
                for line in html.strip().splitlines():
                    u = line.strip()
                    if u.startswith('http') and self.same_domain(u, domain):
                        all_urls.append(u)
                    if len(all_urls) >= MAX_SITEMAP_URLS:
                        break
                break

        return all_urls

    # ── GOOGLE SITE: SEARCH FALLBACK ───────────────────────────────────────────────

    def google_site_search(self, domain, max_results=50):
        """
        Scrape Google search for site:domain to discover indexed pages.
        Returns list of page URLs.
        """
        urls = []
        start = 0

        while len(urls) < max_results:
            query = f'site:{domain}'
            search_url = (
                f'https://www.google.com/search?q={requests.utils.quote(query)}'
                f'&num=10&start={start}&hl=en'
            )
            html, _, code = self.fetch(search_url)
            if not html or code != 200:
                break

            try:
                soup = BeautifulSoup(html, 'html.parser')
            except Exception as e:
                log.warning(f"Google Search parsing failed: {e}")
                break

            # Extract result URLs from Google's cite/a tags
            found_this_page = 0
            for a in soup.select('a[href]'):
                href = a.get('href', '')
                # Google wraps results as /url?q=...
                if '/url?q=' in href:
                    actual = href.split('/url?q=')[1].split('&')[0]
                    actual = requests.utils.unquote(actual)
                    if self.same_domain(actual, domain) and actual not in urls:
                        urls.append(actual)
                        found_this_page += 1

            # Also try direct result anchors
            for div in soup.select('div.g a[href]'):
                href = div.get('href', '')
                if href.startswith('http') and self.same_domain(href, domain):
                    if href not in urls:
                        urls.append(href)
                        found_this_page += 1

            if found_this_page == 0:
                break

            start += 10
            time.sleep(1.5)   # respectful delay for Google

        return urls[:max_results]

    # ── MATCH ENGINE ───────────────────────────────────────────────────────────────

    @staticmethod
    def extract_sentence(text, term, context_chars=150):
        """Extract the sentence around a matched term."""
        lo = text.lower()
        pos = lo.find(term.lower())
        if pos == -1:
            return ''
        # Walk back to sentence start
        start = max(0, pos - context_chars)
        end   = min(len(text), pos + len(term) + context_chars)
        snippet = text[start:end].strip()
        # Clean whitespace
        snippet = re.sub(r'\s+', ' ', snippet)
        return f'...{snippet}...'

    def check_page(self, html, page_url):
        """
        Check a fetched page for matches.
        Returns list of match dicts (may be empty).
        Each dict: { match_type, matched_value, snippet }
        """
        if not html:
            return []

        try:
            soup = BeautifulSoup(html, 'html.parser')
        except Exception as e:
            log.warning(f"Failed to parse page {page_url}: {e}")
            return []

        matches = []

        # ── Check 1: TEXT match — "bajaj allianz life" in visible text ──
        text = soup.get_text(separator=' ', strip=True)
        if TEXT_MATCH.lower() in text.lower():
            snippet = self.extract_sentence(text, TEXT_MATCH)
            matches.append({
                'match_type'   : 'Text',
                'matched_value': TEXT_MATCH,
                'snippet'      : snippet,
            })

        # ── Check 2: HREF match — "bajajallianzlife.com" in any <a href> ──
        for a_tag in soup.find_all('a', href=True):
            href = a_tag.get('href', '').lower()
            if HREF_MATCH.lower() in href:
                anchor_text = a_tag.get_text(strip=True)
                matches.append({
                    'match_type'   : 'Href Link',
                    'matched_value': a_tag['href'],
                    'snippet'      : f'Link text: "{anchor_text}" → href: {a_tag["href"]}',
                })

        # If both found, mark as Both
        types = {m['match_type'] for m in matches}
        if 'Text' in types and 'Href Link' in types:
            for m in matches:
                m['match_type'] = 'Both (Text + Href)'

        return matches

    def extract_internal_links(self, html, base, domain):
        """Extract all internal links from a page."""
        if not html:
            return []
            
        try:
            soup  = BeautifulSoup(html, 'html.parser')
        except Exception as e:
            log.warning(f"Failed to parse links on {base}: {e}")
            return []
            
        links = set()
        for a in soup.find_all('a', href=True):
            if not a.has_attr('href'):
                continue
            href = a['href'].strip()
            # Skip anchors, js, mailto
            if href.startswith(('#', 'javascript:', 'mailto:', 'tel:')):
                continue
            full = urljoin(base, href)
            # Strip fragment
            full = full.split('#')[0].rstrip('/')
            if self.same_domain(full, domain) and full.startswith('http'):
                links.add(full)
        return list(links)

    # ── MAIN CRAWL ENGINE ──────────────────────────────────────────────────────────

    def crawl_domain(self, domain, agent_name):
        """
        Full crawl of one domain.
        Returns list of match dicts.
        """
        log.info(f'  [{domain}] Starting crawl...')
        all_matches   = []
        visited       = set()

        # ── Step 1: Seed URLs ──────────────────────────────────────────────────────
        seed_urls = []

        # Try sitemap first
        sitemap_urls = self.get_sitemap_urls(domain)
        if sitemap_urls:
            log.info(f'  [{domain}] Sitemap: {len(sitemap_urls)} URLs')
            seed_urls = sitemap_urls
        else:
            # Fallback: Google site: search
            log.info(f'  [{domain}] No sitemap — trying Google site: search')
            google_urls = self.google_site_search(domain, max_results=50)
            if google_urls:
                log.info(f'  [{domain}] Google site: search: {len(google_urls)} URLs')
                seed_urls = google_urls
            else:
                log.info(f'  [{domain}] No Google results — crawling from homepage only')

        # Always include homepage
        homepage = self.base_url(domain)
        if homepage not in seed_urls:
            seed_urls.insert(0, homepage)

        # ── Step 2: BFS crawl up to MAX_DEPTH ─────────────────────────────────────
        # Queue items: (url, depth)
        queue   = deque()
        for u in seed_urls:
            queue.append((u, 0))

        pages_checked = 0

        while queue and pages_checked < MAX_PAGES_PER_SITE:
            url, depth = queue.popleft()

            # Normalise URL
            url = url.split('#')[0].rstrip('/')
            if not url or url in visited:
                continue
            if not self.same_domain(url, domain):
                continue

            visited.add(url)
            pages_checked += 1

            # Fetch
            html, final_url, code = self.fetch(url)
            time.sleep(DELAY_BETWEEN_REQ)

            if not html or code not in (200, 301, 302):
                continue

            # Check for matches
            page_matches = self.check_page(html, final_url)
            for m in page_matches:
                all_matches.append({
                    'Agent Name'   : agent_name,
                    'Domain'       : domain,
                    'Page URL'     : final_url,
                    'Depth'        : depth,
                    'Match Type'   : m['match_type'],
                    'Matched Value': m['matched_value'],
                    'Snippet'      : m['snippet'],
                    'Checked At'   : datetime.now().strftime('%Y-%m-%d %H:%M'),
                })
                log.info(f'    ✅ MATCH [{m["match_type"]}] {final_url}')

            # Enqueue internal links if within depth limit
            if depth < MAX_DEPTH:
                internal = self.extract_internal_links(html, final_url, domain)
                for link in internal:
                    link = link.rstrip('/')
                    if link not in visited:
                        queue.append((link, depth + 1))

        log.info(f'  [{domain}] Done — {pages_checked} pages, {len(all_matches)} matches')
        return all_matches, pages_checked

    # ── EXCEL OUTPUT ───────────────────────────────────────────────────────────────

    def build_excel(self, results_df, summary_df):
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        C_NAVY = '1B2A4A'; C_GD = '1A7340'; C_GL = 'E6F4EA'
        C_BLUE = '1967D2'; C_RED = 'B71C1C'; C_OR = 'E65100'
        C_GRY  = '5F6368'; C_GRL = 'F5F5F5'; C_ALT = 'F8F9FA'; WHT = 'FFFFFF'

        thin = Side(style='thin', color='D8D8D8')
        bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)

        def H(c, bg=C_NAVY, fg=WHT, sz=9):
            c.font      = Font(name='Arial', bold=True, color=fg, size=sz)
            c.fill      = PatternFill('solid', fgColor=bg)
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            c.border    = bdr

        def D(c, bold=False, fg='333333', bg=WHT, wrap=False, align='left', lnk=False, sz=9):
            c.font      = Font(name='Arial', bold=bold, color=fg, size=sz,
                               underline='single' if lnk else None)
            c.fill      = PatternFill('solid', fgColor=bg)
            c.alignment = Alignment(horizontal=align, vertical='center', wrap_text=wrap)
            c.border    = bdr

        wb = Workbook()

        # ── Sheet 1: Summary ──────────────────────────────────────────────────────
        ws = wb.active
        ws.title = 'Summary'

        total_domains  = len(summary_df)
        found_domains  = (summary_df['Matches Found'] > 0).sum()
        total_matches  = len(results_df)
        text_matches   = (results_df['Match Type'].str.contains('Text', na=False)).sum()
        href_matches   = (results_df['Match Type'].str.contains('Href', na=False)).sum()

        ws.merge_cells('A1:H2')
        c = ws['A1']
        c.value     = 'Bajaj Allianz Life — Deep Partner Website Audit'
        c.font      = Font(name='Arial', bold=True, size=14, color=WHT)
        c.fill      = PatternFill('solid', fgColor=C_NAVY)
        c.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 18
        ws.row_dimensions[2].height = 18

        ws.merge_cells('A3:H3')
        s = ws['A3']
        s.value     = (f'Source: {os.path.basename(self.input_file)}  |  '
                       f'Crawl: sitemap + Google site: fallback + 3-level recursive  |  '
                       f'Run date: {datetime.now().strftime("%d %b %Y")}')
        s.font      = Font(name='Arial', italic=True, size=9, color='555555')
        s.fill      = PatternFill('solid', fgColor='EEF2F7')
        s.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[3].height = 15
        ws.row_dimensions[4].height = 8

        # KPI cards
        kpis = [
            ('A', 'B', 'TOTAL DOMAINS',     total_domains,  C_NAVY),
            ('C', 'D', 'DOMAINS WITH MATCH',found_domains,  C_GD),
            ('E', 'F', 'TOTAL PAGES FLAGGED',total_matches, C_OR),
            ('G', 'H', 'TEXT MATCHES',       text_matches,  '1565C0'),
        ]
        for ca, cb, lbl, val, bg in kpis:
            ws.merge_cells(f'{ca}5:{cb}5')
            ws.merge_cells(f'{ca}6:{cb}6')
            for cell, v, sz in [(ws[f'{ca}5'], lbl, 8), (ws[f'{ca}6'], val, 22)]:
                cell.value      = v
                cell.font       = Font(name='Arial', bold=True, color=WHT, size=sz)
                cell.fill       = PatternFill('solid', fgColor=bg)
                cell.alignment  = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[5].height = 18
        ws.row_dimensions[6].height = 34

        ws.merge_cells('A7:H7')
        hc = ws['A7']
        hc.value     = (f'HREF matches (bajajallianzlife.com in link): {href_matches}  |  '
                        f'Both text + href: {(results_df["Match Type"].str.contains("Both", na=False)).sum()}')
        hc.font      = Font(name='Arial', bold=True, size=9, color=C_BLUE)
        hc.fill      = PatternFill('solid', fgColor='E8F0FE')
        hc.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[7].height = 20
        ws.row_dimensions[8].height = 10

        # Domain summary table
        r = 9
        ws.cell(row=r, column=1).value = 'Per-Domain Summary'
        ws.cell(row=r, column=1).font  = Font(name='Arial', bold=True, size=10, color=C_NAVY)
        ws.row_dimensions[r].height    = 20
        r += 1

        sum_cols = ['Domain', 'Agent Name', 'Pages Crawled', 'Matches Found', 'Match Types', 'Crawl Status']
        for ci, h in enumerate(sum_cols, 1):
            H(ws.cell(row=r, column=ci, value=h), bg='2C3E6B')
        ws.row_dimensions[r].height = 22
        r += 1

        for _, row in summary_df.sort_values('Matches Found', ascending=False).iterrows():
            bg      = C_GL if row['Matches Found'] > 0 else (C_GRL if r % 2 == 0 else WHT)
            has_match = row['Matches Found'] > 0
            for ci, field in enumerate(sum_cols, 1):
                v = str(row.get(field, ''))
                v = '' if v in ('nan', 'None') else v
                c = ws.cell(row=r, column=ci, value=v)
                D(c, bold=(ci == 4 and has_match),
                  fg=C_GD if has_match else '333333', bg=bg,
                  align='center' if ci in (3, 4) else 'left')
            ws.row_dimensions[r].height = 18
            r += 1

        for i, w in enumerate([28, 36, 14, 14, 28, 20], 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        # ── Sheet 2: All Matches (one row per matched page) ───────────────────────
        ws2  = wb.create_sheet('✅ All Matches')
        cols = [
            ('Sr',          5),
            ('Domain',      26),
            ('Agent Name',  34),
            ('Page URL',    58),
            ('Depth',        7),
            ('Match Type',  20),
            ('Matched Value',36),
            ('Snippet / Context', 70),
            ('Checked At',  18),
        ]
        for ci, (h, w) in enumerate(cols, 1):
            H(ws2.cell(row=1, column=ci, value=h))
            ws2.column_dimensions[get_column_letter(ci)].width = w
        ws2.row_dimensions[1].height = 26
        ws2.freeze_panes = 'A2'

        for ri, (_, row) in enumerate(results_df.iterrows(), 2):
            mtype = str(row.get('Match Type', ''))
            bg    = C_GL if 'Text' in mtype else ('E8F0FE' if 'Href' in mtype else C_ALT)
            if ri % 2 == 0 and bg == C_ALT:
                bg = 'F0F4FF'

            for ci, (h, _) in enumerate(cols, 1):
                raw = row.get(h, '')
                v   = '' if str(raw) in ('nan', 'None') else str(raw)
                c   = ws2.cell(row=ri, column=ci, value=v)

                if h == 'Match Type':
                    if 'Both' in v:   fc = '6A1B9A'; cbg = 'F3E5F5'
                    elif 'Text' in v: fc = C_GD;     cbg = C_GL
                    else:             fc = C_BLUE;   cbg = 'E8F0FE'
                    c.value     = v
                    c.font      = Font(name='Arial', bold=True, color=fc, size=9)
                    c.fill      = PatternFill('solid', fgColor=cbg)
                    c.alignment = Alignment(horizontal='center', vertical='center')
                    c.border    = bdr
                elif h == 'Page URL' and v.startswith('http'):
                    D(c, fg=C_BLUE, bg=bg, lnk=True)
                elif h == 'Matched Value' and 'bajajallianzlife' in v.lower():
                    D(c, bold=True, fg=C_RED, bg=bg)
                elif h == 'Snippet / Context':
                    D(c, bg=bg, wrap=True, fg='444444')
                else:
                    D(c, bg=bg)

            ws2.row_dimensions[ri].height = 32 if row.get('Snippet / Context') else 18

        ws2.auto_filter.ref = f'A1:{get_column_letter(len(cols))}1'

        # ── Sheet 3: Text Matches only ────────────────────────────────────────────
        ws3     = wb.create_sheet('📄 Text Matches')
        text_df = results_df[results_df['Match Type'].str.contains('Text', na=False)]
        self._build_filtered_sheet(ws3, text_df, 'Text Matches — "bajaj allianz life" in page content',
                                   C_GD, C_GL, cols, bdr, H, D, C_BLUE, C_GD)

        # ── Sheet 4: Href Matches only ────────────────────────────────────────────
        ws4     = wb.create_sheet('🔗 Href Matches')
        href_df = results_df[results_df['Match Type'].str.contains('Href', na=False)]
        self._build_filtered_sheet(ws4, href_df, 'Href Matches — "bajajallianzlife.com" found in page links',
                                   '1565C0', 'E8F0FE', cols, bdr, H, D, C_BLUE, C_GD)

        wb.save(self.output_excel)
        log.info(f'Excel saved: {self.output_excel}')

    def _build_filtered_sheet(self, ws, df, title, hdr_bg, row_bg, cols, bdr, H, D, C_BLUE, C_GD):
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        WHT = 'FFFFFF'

        ws.merge_cells(f'A1:{get_column_letter(len(cols))}1')
        c = ws['A1']
        c.value     = title
        c.font      = Font(name='Arial', bold=True, size=11, color=WHT)
        c.fill      = PatternFill('solid', fgColor=hdr_bg)
        c.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 26

        for ci, (h, w) in enumerate(cols, 1):
            H(ws.cell(row=2, column=ci, value=h), bg=hdr_bg)
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.row_dimensions[2].height = 22
        ws.freeze_panes = 'A3'

        for ri, (_, row) in enumerate(df.iterrows(), 3):
            bg = row_bg if ri % 2 == 0 else WHT
            for ci, (h, _) in enumerate(cols, 1):
                raw = row.get(h, '')
                v   = '' if str(raw) in ('nan', 'None') else str(raw)
                c   = ws.cell(row=ri, column=ci, value=v)
                if h == 'Page URL' and v.startswith('http'):
                    D(c, fg=C_BLUE, bg=bg, lnk=True)
                elif h == 'Snippet / Context':
                    D(c, bg=bg, wrap=True, fg='444444')
                elif h == 'Match Type':
                    D(c, bold=True, fg=hdr_bg, bg=bg, align='center')
                else:
                    D(c, bg=bg)
            ws.row_dimensions[ri].height = 32

        ws.auto_filter.ref = f'A2:{get_column_letter(len(cols))}2'

        if len(df) == 0:
            ws.cell(row=3, column=1).value = 'No matches of this type found.'
            ws.cell(row=3, column=1).font  = Font(name='Arial', italic=True,
                                                   size=10, color='999999')

    def run_audit(self):
        log.info('=' * 70)
        log.info('  Bajaj Allianz Life Deep Partner Audit — Starting')
        log.info('=' * 70)

        # Load domain list
        try:
            df_sheet4 = pd.read_excel(self.input_file, sheet_name='Sheet4')
            df_sheet1 = pd.read_excel(self.input_file, sheet_name='Sheet1')
        except FileNotFoundError:
            log.error(f'Could not find input file: {self.input_file}')
            return
        except Exception as e:
            log.error(f'Error reading Excel file: {e}')
            return

        # Build agent name lookup: domain → agent name
        agent_lookup = {}
        for _, row in df_sheet1.iterrows():
            link = str(row.get('Links', '')).strip()
            if link.startswith('http'):
                try:
                    domain = urlparse(link).netloc.replace('www.', '').lower()
                    agent_lookup[domain] = str(row.get('AGENT_NAME', ''))
                except:
                    pass

        # Filter to valid domains
        valid_mask = df_sheet4['Clean Domain'].str.contains(r'\.', regex=True, na=False)
        domains_df = df_sheet4[valid_mask].copy().reset_index(drop=True)
        log.info(f'Domains to crawl: {len(domains_df)}')

        # Load progress if resuming
        done_domains = set()
        all_results  = []
        all_summary  = []

        if os.path.exists(self.progress_csv):
            prog = pd.read_csv(self.progress_csv, dtype=str).fillna('')
            done_domains = set(prog['Domain'].unique())
            all_results  = prog.to_dict('records')
            log.info(f'Resuming — {len(done_domains)} domains already done')

        total = len(domains_df)

        progress_lock = threading.Lock()
        processed_count = 0

        def process_domain(idx, row):
            domain = str(row['Clean Domain']).strip().lower()
            domain = domain.replace('https://', '').replace('http://', '').strip('/')

            if domain in done_domains:
                return None

            agent_name = agent_lookup.get(domain, '')
            if not agent_name:
                # Try partial match
                for k, v in agent_lookup.items():
                    if k in domain or domain in k:
                        agent_name = v
                        break

            log.info(f'[{idx+1}/{total}] Starting thread for {domain}  ({agent_name[:40]})')

            try:
                matches, pages_checked = self.crawl_domain(domain, agent_name)
            except Exception as e:
                log.error(f'  ERROR on {domain}: {e}')
                matches       = []
                pages_checked = 0
                
            return {
                'domain': domain,
                'agent_name': agent_name,
                'matches': matches,
                'pages_checked': pages_checked
            }

        with concurrent.futures.ThreadPoolExecutor(max_workers=10) as executor:
            # Submit all domains
            future_to_domain = {
                executor.submit(process_domain, idx, row): (idx, row)
                for idx, row in domains_df.iterrows()
            }

            for future in concurrent.futures.as_completed(future_to_domain):
                result = future.result()
                if not result:
                    continue
                    
                # Thread-safe update and save
                with progress_lock:
                    all_results.extend(result['matches'])
                    
                    match_types = list({m['Match Type'] for m in result['matches']})
                    all_summary.append({
                        'Domain'       : result['domain'],
                        'Agent Name'   : result['agent_name'],
                        'Pages Crawled': result['pages_checked'],
                        'Matches Found': len(result['matches']),
                        'Match Types'  : ' | '.join(match_types) if match_types else '-',
                        'Crawl Status' : 'Done',
                    })
                    
                    done_domains.add(result['domain'])
                    processed_count += 1

                    # Save progress safely
                    if processed_count % 5 == 0:
                        pd.DataFrame(all_results).to_csv(self.progress_csv, index=False)
                        log.info(f'  Progress saved ({len(done_domains)}/{total} done, '
                                 f'{len(all_results)} matches so far)')

        # Final save
        results_df = pd.DataFrame(all_results) if all_results else pd.DataFrame(columns=[
            'Agent Name','Domain','Page URL','Depth','Match Type',
            'Matched Value','Snippet','Checked At'])
        results_df.rename(columns={'Snippet': 'Snippet / Context'}, inplace=True)

        summary_df = pd.DataFrame(all_summary) if all_summary else pd.DataFrame(columns=[
            'Domain','Agent Name','Pages Crawled','Matches Found','Match Types','Crawl Status'])

        log.info('=' * 70)
        log.info(f'CRAWL COMPLETE')
        log.info(f'  Domains crawled : {len(all_summary)}')
        log.info(f'  Domains with matches: {(summary_df["Matches Found"] > 0).sum()}')
        log.info(f'  Total matched pages : {len(results_df)}')
        log.info('=' * 70)

        self.build_excel(results_df, summary_df)

        # Clean up progress file
        if os.path.exists(self.progress_csv):
            os.remove(self.progress_csv)

        print(f'\n✅ Done! Output: {self.output_excel}')
        print(f'   Matched pages: {len(results_df)}')
        print(f'   Domains flagged: {(summary_df["Matches Found"] > 0).sum()} / {len(summary_df)}')
